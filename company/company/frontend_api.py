import frappe
from frappe.auth import LoginManager
from frappe import _
from livekit import api
from frappe.utils import getdate, add_months, get_first_day, get_last_day, flt, today
from datetime import datetime


@frappe.whitelist(allow_guest=True)
def get_csrf_token():
    return frappe.sessions.get_csrf_token()

@frappe.whitelist()
def get_invoice_count():
    return frappe.db.count("Invoice")

@frappe.whitelist()
def get_doctype_list(doctype, txt=None, fields=None, filters=None):
    """
    Fetch a list of documents for a given DocType.
    Useful for populating dropdowns on the frontend.
    """
    if doctype == "Payment Terms":
        # Bypass direct permission check for Payment Terms to allow Invoice readers/creators to fetch them
        if not frappe.has_permission("Invoice", "read") and not frappe.has_permission("Invoice", "create"):
            return []
    elif not frappe.has_permission(doctype, "read"):
        return []

    query_filters = {}
    if txt:
        query_filters["name"] = ["like", f"%{txt}%"]

    if filters:
        import json
        extra_filters = json.loads(filters)
        query_filters.update(extra_filters)

    # Use get_all for Payment Terms to bypass strict system manager check
    fetch_fn = frappe.get_all if doctype == "Payment Terms" else frappe.get_list

    if fields:
        import json
        field_list = json.loads(fields)
        return fetch_fn(doctype, filters=query_filters, fields=field_list, limit=1000)

    return fetch_fn(doctype, filters=query_filters, pluck="name", limit=1000)


@frappe.whitelist()
def get_permitted_count(doctype, filters=None, or_filters=None):
    """
    Returns the count of documents the user is permitted to see.
    Standard frappe.client.get_count is not always permission-aware.
    """
    import json
    if isinstance(filters, str):
        filters = json.loads(filters)
    if isinstance(or_filters, str):
        or_filters = json.loads(or_filters)

    if doctype == "Contacts" and or_filters:
        or_filters = clean_contacts_or_filters(or_filters)

    if doctype in ("Leave Application", "Request", "WFH Attendance", "Reimbursement Claim", "Asset Request") and filters:
        if isinstance(filters, list):
            filters = [f for f in filters if not (isinstance(f, list) and len(f) >= 3 and (f[1] if len(f) == 4 else f[0]) in ("unread_only", "unread_messages"))]
        elif isinstance(filters, dict):
            filters.pop("unread_only", None)
            filters.pop("unread_messages", None)


    # Resolve company_name filter for Contacts (child table field)
    if doctype == "Contacts" and filters:
        new_filters = []
        for f in filters:
            if isinstance(f, list) and len(f) >= 3:
                fieldname = f[1] if len(f) == 4 else f[0]
                if fieldname == "company_name":
                    company_val = f[3] if len(f) == 4 else f[2]
                    matching_account = frappe.db.get_value("Accounts", {"account_name": company_val}, "name")
                    if matching_account:
                        matching_parents = frappe.get_all(
                            "Contact Company",
                            filters={"company_name": matching_account, "parenttype": "Contacts"},
                            pluck="parent"
                        )
                    else:
                        matching_parents = []
                    new_filters.append(["Contacts", "name", "in", matching_parents if matching_parents else [""]])
                else:
                    new_filters.append(f)
            else:
                new_filters.append(f)
        filters = new_filters

    return len(frappe.get_list(
        doctype,
        filters=filters,
        or_filters=or_filters,
        pluck="name",
        limit=None
    ))


@frappe.whitelist()
def check_asset_availability(asset, name=None):
    """
    Check if the Asset is already assigned and not yet returned.
    Returns details of the existing assignment if found.
    """
    filters = {
        "asset": asset,
        "returned_on": ["is", "not set"]
    }
    if name:
        filters["name"] = ["!=", name]

    existing = frappe.db.get_value(
        "Asset Assignment",
        filters,
        ["name", "assigned_to", "employee_name"],
        as_dict=True
    )

    if existing:
        return {
            "is_assigned": True,
            "assigned_to": existing.assigned_to,
            "employee_name": existing.employee_name,
            "assignment_name": existing.name
        }

    return {"is_assigned": False}


@frappe.whitelist(allow_guest=True)
def mobile_login(username, password):
    """
    Login API for Mobile / App.
    Returns API Key + Secret per user.

    - If the user already has api_key + api_secret → returns existing ones (no regeneration).
    - If either is missing → calls Frappe's built-in generate_keys to create them once.

    This ensures existing mobile sessions are never broken by a login call.
    """
    from frappe.core.doctype.user.user import generate_keys

    if not username or not password:
        return {"status": "failed", "message": "Username and password required"}

    # --- Authenticate ---
    try:
        login_manager = LoginManager()
        login_manager.authenticate(username, password)
        login_manager.post_login()
    except frappe.AuthenticationError:
        return {"status": "failed", "message": "Invalid login credentials"}

    # Prevent Frappe from appending home_page and full_name to the root of the API response
    frappe.local.response.pop('home_page', None)
    frappe.local.response.pop('full_name', None)

    logged_in_user = frappe.session.user
    user = frappe.get_doc("User", logged_in_user)

    # --- Return existing keys if already set (do NOT regenerate) ---
    if user.api_key and user.api_secret:
        return {
            "status": "success",
            "message": "Login successful",
            "data": {
                "user":       logged_in_user,
                "api_key":    user.api_key,
                "api_secret": user.get_password("api_secret"),  # decrypted plain-text
            }
        }

    # --- Keys are missing — generate them once via Frappe's built-in generate_keys ---
    # generate_keys requires System Manager role, so we temporarily
    # elevate to Administrator to make the call, then restore the session.
    frappe.set_user("Administrator")
    try:
        keys = generate_keys(logged_in_user)
    finally:
        frappe.set_user(logged_in_user)

    frappe.db.commit()

    return {
        "user":       logged_in_user,
        "api_key":    keys.get("api_key"),
        "api_secret": keys.get("api_secret"),   # plain-text — store securely on device
    }


@frappe.whitelist(allow_guest=False)
def mobile_get_announcements(limit=20):
    """
    Fetch active Announcements for HRMS Mobile App.
    Only returns records where is_active = 1, newest first.

    Requires Token Auth: Authorization: token <api_key>:<api_secret>

    Args:
        limit (int): Max number of announcements to return. Default 20.

    Returns:
        list of { name, announcement_name, announcement, is_active, creation }
    """
    try:
        announcements = frappe.get_all(
            "Announcement",
            filters={"is_active": 1},
            fields=["name", "announcement_name", "announcement", "is_active", "creation"],
            order_by="creation desc",
            limit=int(limit)
        )
        return announcements
    except Exception as e:
        frappe.log_error(f"Error fetching announcements: {str(e)}")
        return []


@frappe.whitelist()
def get_lead_permissions():
    """
    Check if the current user has read, write, and delete permissions for the Lead DocType.
    """
    return get_doc_permissions("Lead")


@frappe.whitelist()
def get_deal_permissions():
    """
    Check if the current user has read, write, and delete permissions for the Deal DocType.
    """
    return get_doc_permissions("Deal")


@frappe.whitelist()
def get_doc_permissions(doctype):
    """
    Check if the current user has read, write, and delete permissions for a given DocType.
    """
    return {
        "read": bool(frappe.has_permission(doctype, "read")),
        "write": bool(frappe.has_permission(doctype, "write")),
        "create": bool(frappe.has_permission(doctype, "create")),
        "delete": bool(frappe.has_permission(doctype, "delete")),
    }


@frappe.whitelist()
def get_current_user_info():
    """
    Fetch the full details of the currently logged-in user.
    """
    user = frappe.get_doc("User", frappe.session.user)

    # Calculate allowed modules (All - Blocked)
    all_modules = frappe.get_all("Module Def", pluck="name")
    blocked_modules = [d.module for d in user.block_modules]
    allowed_modules = [m for m in all_modules if m not in blocked_modules]

    # Fetch employee info
    employee = frappe.db.get_value("Employee", {"user": user.name}, ["name", "employee_name"], as_dict=True)

    has_crm_permission = bool(frappe.db.exists("User Permission", {"user": user.name}))

    return {
        "status": "success",
        "message": "User info fetched successfully",
        "data": {
            "name": user.name,
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "full_name": user.full_name,
            "username": user.username,
            "email": user.email,
            "time_zone": user.time_zone,
            "user_image": user.user_image,
            "roles": [role.role for role in user.roles],
            "role_profile_name": user.role_profile_name,
            "allowed_modules": allowed_modules,
            "employee": employee.get("name") if employee else None,
            "employee_name": employee.get("employee_name") if employee else None,
            "has_crm_permission": has_crm_permission
        }
    }


@frappe.whitelist()
def get_dashboard_stats(start_date=None, end_date=None):
    """
    Fetch CRM dashboard statistics including counts for Leads, Contacts, Deals, Events, Todo, Calls, and Meetings.
    """
    stats = {}
    user = frappe.session.user

    # Check if user has a User Permission restricting their view to their own records
    has_user_permission = frappe.db.exists("User Permission", {"user": user, "allow": "User"})

    # Get counts for each DocType
    doctypes = {
        "leads": "Lead",
        "contacts": "Contacts",
        "accounts": "Accounts",
        "deals": "Deal",
        "proposal": "Proposal",
        "estimation": "Estimation",
        "invoice": "Invoice"
    }

    date_filter = {}
    if start_date and end_date:
        date_filter["creation"] = ["between", [start_date, f"{end_date} 23:59:59"]]
    elif start_date:
        date_filter["creation"] = [">=", start_date]
    elif end_date:
        date_filter["creation"] = ["<=", f"{end_date} 23:59:59"]

    for key, doctype in doctypes.items():
        try:
            if frappe.has_permission(doctype, "read"):
                filters = {}
                if has_user_permission:
                    filters["owner_name"] = user
                filters.update(date_filter)
                stats[key] = frappe.db.count(doctype, filters)
            else:
                stats[key] = 0
        except Exception:
            stats[key] = 0

    # Get leads by status (workflow_state)
    try:
        if frappe.has_permission("Lead", "read"):
            cond = ""
            if start_date and end_date:
                cond = f"AND DATE(creation) BETWEEN '{start_date}' AND '{end_date}'"
            if has_user_permission:
                stats["leads_by_status"] = frappe.db.sql(f"""
                    SELECT workflow_state as status, COUNT(*) as count
                    FROM `tabLead`
                    WHERE owner_name = %s {cond}
                    GROUP BY workflow_state
                """, (user,), as_dict=True)
            else:
                stats["leads_by_status"] = frappe.db.sql(f"""
                    SELECT workflow_state as status, COUNT(*) as count
                    FROM `tabLead`
                    WHERE 1=1 {cond}
                    GROUP BY workflow_state
                """, as_dict=True)
        else:
            stats["leads_by_status"] = []
    except Exception:
        stats["leads_by_status"] = []

    # Get deals by stage
    try:
        if frappe.has_permission("Deal", "read"):
            cond = ""
            if start_date and end_date:
                cond = f"AND DATE(creation) BETWEEN '{start_date}' AND '{end_date}'"
            if has_user_permission:
                stats["deals_by_stage"] = frappe.db.sql(f"""
                    SELECT stage, COUNT(*) as count
                    FROM `tabDeal`
                    WHERE owner_name = %s {cond}
                    GROUP BY stage
                """, (user,), as_dict=True)
            else:
                stats["deals_by_stage"] = frappe.db.sql(f"""
                    SELECT stage, COUNT(*) as count
                    FROM `tabDeal`
                    WHERE 1=1 {cond}
                    GROUP BY stage
                """, as_dict=True)
        else:
            stats["deals_by_stage"] = []
    except Exception:
        stats["deals_by_stage"] = []

    # Get historical data for the last 7 days
    try:
        days = []
        lead_series = []
        contact_series = []
        account_series = []
        deal_series = []
        proposal_series = []
        estimation_series = []
        invoice_series = []

        for i in range(6, -1, -1):
            date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
            day_name = frappe.utils.get_datetime(date).strftime('%a')
            days.append(day_name)

            if has_user_permission:
                lead_series.append(frappe.db.count("Lead", {"creation": ["like", f"{date}%"], "owner_name": user}))
                contact_series.append(frappe.db.count("Contacts", {"creation": ["like", f"{date}%"], "owner_name": user}))
                account_series.append(frappe.db.count("Accounts", {"creation": ["like", f"{date}%"], "owner_name": user}))
                deal_series.append(frappe.db.count("Deal", {"creation": ["like", f"{date}%"], "owner_name": user}))
                proposal_series.append(frappe.db.count("Proposal", {"creation": ["like", f"{date}%"], "owner_name": user}))
                estimation_series.append(frappe.db.count("Estimation", {"creation": ["like", f"{date}%"], "owner_name": user}))
                invoice_series.append(frappe.db.count("Invoice", {"creation": ["like", f"{date}%"], "owner_name": user}))
            else:
                lead_series.append(frappe.db.count("Lead", {"creation": ["like", f"{date}%"]}))
                contact_series.append(frappe.db.count("Contacts", {"creation": ["like", f"{date}%"]}))
                account_series.append(frappe.db.count("Accounts", {"creation": ["like", f"{date}%"]}))
                deal_series.append(frappe.db.count("Deal", {"creation": ["like", f"{date}%"]}))
                proposal_series.append(frappe.db.count("Proposal", {"creation": ["like", f"{date}%"]}))
                estimation_series.append(frappe.db.count("Estimation", {"creation": ["like", f"{date}%"]}))
                invoice_series.append(frappe.db.count("Invoice", {"creation": ["like", f"{date}%"]}))

        stats["charts"] = {
            "categories": days,
            "leads": lead_series,
            "contacts": contact_series,
            "accounts": account_series,
            "deals": deal_series,
            "proposals": proposal_series,
            "estimations": estimation_series,
            "invoices": invoice_series
        }
    except Exception as e:
        frappe.log_error(f"Error calculating dashboard chart data: {str(e)}")
        stats["charts"] = {
            "categories": ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
            "leads": [0]*7,
            "contacts": [0]*7,
            "deals": [0]*7,
            "accounts": [0]*7
        }

    return stats


@frappe.whitelist()
def get_expense_tracker_stats(start_date=None, end_date=None):
    """
    Get dashboard stats for Expense Tracker.
    Calculates total income, total expense and balance based on the period.
    """
    filters = {}
    if start_date and end_date:
        filters["date_time"] = ["between", [start_date, end_date]]
    elif start_date:
        filters["date_time"] = [">=", start_date]
    elif end_date:
        filters["date_time"] = ["<=", end_date]

    stats = {
        "total_income": 0,
        "total_expense": 0,
        "balance": 0
    }

    try:
        data = frappe.get_all("Expense Tracker", filters=filters, fields=["type", "amount"])

        for d in data:
            if d.type == "Income":
                stats["total_income"] += frappe.utils.flt(d.amount)
            elif d.type == "Expense":
                stats["total_expense"] += frappe.utils.flt(d.amount)

        stats["balance"] = stats["total_income"] - stats["total_expense"]
    except Exception as e:
        frappe.log_error(f"Error fetching Expense Tracker stats: {str(e)}")

    return stats

@frappe.whitelist()
def update_request_status(name, workflow_state, update_data=None):
    """
    Update the workflow state of a Request.
    Handles doc_status change for 'Rejected' state.
    """
    if isinstance(update_data, str):
        import json
        update_data = json.loads(update_data)

    doc = frappe.get_doc("Request", name)

    if update_data:
        doc.update(update_data)

    doc.workflow_state = workflow_state

    if workflow_state == "Rejected":
        doc.cancel()
    else:
        doc.save()

    # Push real-time update so Request dialogs refresh instantly
    frappe.publish_realtime(
        event="request_updated",
        message={"name": name, "workflow_state": workflow_state},
    )

    return doc.as_dict()
 
 
@frappe.whitelist()
def update_doc_status(doctype, name, workflow_state, update_data=None):
    """
    Update the workflow state of any document along with optional field updates.
    Uses frappe.db.set_value to bypass workflow validation, timestamp checks,
    and permission checks — this is a privileged internal API.
    """
    if isinstance(update_data, str):
        import json
        update_data = json.loads(update_data)

    # Build the fields dict for direct DB update
    fields = {"workflow_state": workflow_state}
    if update_data:
        fields.update(update_data)

    # Direct SQL update — bypasses all Frappe validations:
    # WorkflowPermissionError, TimestampMismatchError, permission checks
    frappe.db.set_value(doctype, name, fields)
    frappe.db.commit()

    # Return the refreshed document
    doc = frappe.get_doc(doctype, name)

    # Push real-time update to all connected clients so they can refresh
    if doctype == "Leave Application":
        frappe.publish_realtime(
            event="leave_application_updated",
            message={"name": name, "workflow_state": workflow_state},
        )
    elif doctype == "Asset Request":
        frappe.publish_realtime(
            event="asset_request_updated",
            message={"name": name, "workflow_state": workflow_state},
        )

    return doc.as_dict()


@frappe.whitelist()
def rename_doc(doctype, old, new, merge=False):
    """
    Rename a document. 
    Wrapper for frappe.rename_doc with permission checks.
    """
    if not frappe.has_permission(doctype, "write"):
        frappe.throw(_("Not permitted to rename {0}").format(doctype), frappe.PermissionError)

    return frappe.rename_doc(doctype, old, new, merge=merge)


@frappe.whitelist()
def get_today_activities():
    """
    Fetch today's calls and meetings.
    """
    user = frappe.get_value("User", frappe.session.user, "name")
    from datetime import datetime

    # Only filter by owner if the user has a User Permission record
    has_user_permission = frappe.db.exists("User Permission", {"user": user, "allow": "User"})

    activities = {
        "calls": [],
        "meetings": []
    }

    # Get today's date
    today_date = frappe.utils.today()

    # Fetch and filter calls
    try:
        if frappe.has_permission("Calls", "read"):
            if has_user_permission:
                activities["calls"] = frappe.db.sql("""
                    SELECT name, title, call_for, lead_name, call_start_time, call_end_time, outgoing_call_status, call_purpose
                    FROM `tabCalls`
                    WHERE DATE(call_start_time) = %s
                    AND owner_name = %s
                    ORDER BY call_start_time ASC
                    LIMIT 10
                """, (today_date, user), as_dict=True)
            else:
                activities["calls"] = frappe.db.sql("""
                    SELECT name, title, call_for, lead_name, call_start_time, call_end_time, outgoing_call_status, call_purpose
                    FROM `tabCalls`
                    WHERE DATE(call_start_time) = %s
                    ORDER BY call_start_time ASC
                    LIMIT 10
                """, (today_date,), as_dict=True)
    except Exception as e:
        frappe.log_error(f"Error fetching calls for dashboard: {str(e)}")

    # Fetch and filter meetings (strictly from Meeting DocType)
    try:
        if frappe.has_permission("Meeting", "read"):
            if has_user_permission:
                activities["meetings"] = frappe.db.sql("""
                    SELECT name, title, meet_for, lead_name, `from`, `to`, outgoing_call_status, meeting_venue, location
                    FROM `tabMeeting`
                    WHERE DATE(`from`) = %s
                    AND owner_name = %s
                    ORDER BY `from` ASC
                    LIMIT 10
                """, (today_date, user), as_dict=True)
            else:
                activities["meetings"] = frappe.db.sql("""
                    SELECT name, title, meet_for, lead_name, `from`, `to`, outgoing_call_status, meeting_venue, location
                    FROM `tabMeeting`
                    WHERE DATE(`from`) = %s
                    ORDER BY `from` ASC
                    LIMIT 10
                """, (today_date,), as_dict=True)
    except Exception as e:
        frappe.log_error(f"Error fetching meetings for dashboard: {str(e)}")

    return activities


@frappe.whitelist()
def update_event(name, data):
    """
    Update event data (subject, starts_on, ends_on, etc.)
    """
    if isinstance(data, str):
        import json
        data = json.loads(data)

    try:
        if frappe.has_permission("Event", "write", doc=name):
            doc = frappe.get_doc("Event", name)
            doc.update(data)
            doc.save()
            return {"status": "success", "message": "Event updated successfully"}
        else:
            frappe.throw("No permission to update this event", frappe.PermissionError)
    except Exception as e:
        frappe.log_error(f"Error updating event {name}: {str(e)}")
        return {"status": "error", "message": str(e)}


@frappe.whitelist()
def delete_event(name):
    """
    Delete an event
    """
    try:
        if frappe.has_permission("Event", "delete", doc=name):
            frappe.delete_doc("Event", name)
            return {"status": "success", "message": "Event deleted successfully"}
        else:
            frappe.throw("No permission to delete this event", frappe.PermissionError)
    except Exception as e:
        frappe.log_error(f"Error deleting event {name}: {str(e)}")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def create_event(data):
    """
    Create a new event
    """
    if isinstance(data, str):
        import json
        data = json.loads(data)

    try:
        if frappe.has_permission("Event", "create"):
            doc = frappe.get_doc({
                "doctype": "Event",
                **data
            })
            doc.insert()
            return {"status": "success", "message": "Event created successfully", "name": doc.name}
        else:
            frappe.throw("No permission to create events", frappe.PermissionError)
    except Exception as e:
        frappe.log_error(f"Error creating event: {str(e)}")
        return {"status": "error", "message": str(e)}

@frappe.whitelist()
def get_events(start=None, end=None):
    """
    Fetch events for calendar view.
    """
    filters = {}

    if start and end:
        filters["starts_on"] = ["between", [start, end]]

    try:
        if frappe.has_permission("Event", "read"):
            events = frappe.get_all(
                "Event",
                filters=filters,
                fields=["name", "subject", "starts_on", "ends_on", "event_category", "event_type", "status", "description", "color", "all_day"],
                order_by="starts_on asc",
                limit=500
            )
            return events
        else:
            return []
    except Exception as e:
        frappe.log_error(f"Error fetching events: {str(e)}")
        return []

@frappe.whitelist()
def get_workflow_states(doctype="Lead", current_state=None):
    """Get workflow states and allowed transitions for a doctype and current state"""
    try:
        # Get workflow for the doctype
        workflow = frappe.get_all(
            "Workflow",
            filters={"document_type": doctype, "is_active": 1},
            fields=["name"],
            limit=1
        )

        if not workflow:
            return {"states": [], "transitions": [], "actions": []}

        workflow_name = workflow[0].name

        # Get all workflow states
        states = frappe.get_all(
            "Workflow Document State",
            filters={"parent": workflow_name},
            fields=["state", "doc_status", "is_optional_state"],
            order_by="idx"
        )

        # Get all workflow transitions
        transitions = frappe.get_all(
            "Workflow Transition",
            filters={"parent": workflow_name},
            fields=["state", "action", "next_state", "allowed"],
            order_by="idx"
        )

        # If current_state is provided, filter transitions for that state
        allowed_actions = []
        if current_state:
            user_roles = frappe.get_roles()
            for transition in transitions:
                if transition.state == current_state:
                    # Check if user has the required role
                    if not transition.allowed or transition.allowed in user_roles:
                        allowed_actions.append({
                            "action": transition.action,
                            "next_state": transition.next_state
                        })

        return {
            "states": [s.state for s in states],
            "transitions": transitions,
            "actions": allowed_actions
        }
    except Exception as e:
        frappe.log_error(f"Error fetching workflow states: {str(e)}")
        return {"states": [], "transitions": [], "actions": []}

@frappe.whitelist()
def get_doc_fields(doctype):
    """
    Fetch relevant fields for a given DocType for mapping/import purposes.
    """
    if not frappe.has_permission(doctype, "read"):
        frappe.throw(_("Not permitted"), frappe.PermissionError)

    meta = frappe.get_meta(doctype)
    fields = []

    # Add Name/ID field
    if doctype not in ("Lead", "Contacts", "Accounts"):
        fields.append({
            "fieldname": "name",
            "label": _("ID"),
            "fieldtype": "Data",
            "reqd": 0
        })

    # Add Owner field
    fields.append({
        "fieldname": "owner",
        "label": _("Owner"),
        "fieldtype": "Link",
        "options": "User"
    })

    for df in meta.fields:
        if df.fieldtype not in ("Section Break", "Column Break", "Tab Break", "HTML", "Button"):
            fields.append({
                "fieldname": df.fieldname,
                "label": _(df.label),
                "fieldtype": df.fieldtype,
                "options": df.options,
                "reqd": df.reqd
            })

    return fields


@frappe.whitelist()
def download_import_template(doctype):
    """
    Generate and download a blank import template for a given DocType.
    Customized to force Phone columns as "Text" in Excel.
    """
    import json
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from frappe.core.doctype.data_import.exporter import Exporter
    from frappe.desk.utils import provide_binary_file

    if not frappe.has_permission(doctype, "read"):
        frappe.throw(_("Not permitted"), frappe.PermissionError)

    meta = frappe.get_meta(doctype)
    # Get relevant fields for the template (mandatory + common)
    fields = [df.fieldname for df in meta.fields if (df.reqd or df.in_list_view) and df.fieldtype not in ("Section Break", "Column Break", "Tab Break", "HTML", "Button") and not df.hidden]

    # Ensure requested fields are included for Lead
    if doctype == "Lead":
        extra_fields = ["gstin", "phone_number", "billing_address", "remarks"]
        for f in extra_fields:
            if f not in fields:
                fields.append(f)

    # Ensure requested fields are included for Contacts
    if doctype == "Contacts":
        extra_fields = ["customer_type", "designation", "address", "notes"]
        for f in extra_fields:
            if f not in fields:
                fields.append(f)

    # Ensure requested fields are included for Attendance
    if doctype == "Attendance":
        existing = [f.fieldname if hasattr(f, 'fieldname') else f for f in fields]
        target_fields = ["employee", "employee_name", "attendance_date", "status", "in_time", "out_time"]
        
        new_fields = []
        for fn in target_fields:
            if fn in existing:
                for orig_f in fields:
                    if (orig_f.fieldname if hasattr(orig_f, 'fieldname') else orig_f) == fn:
                        new_fields.append(orig_f)
                        break
            else:
                f_obj = meta.get_field(fn)
                new_fields.append(f_obj if f_obj else fn)
                
        fields = new_fields

    # Ensure requested fields are included for Accounts
    if doctype == "Accounts":
        extra_fields = ["gstin", "website"]
        for f in extra_fields:
            if f not in fields:
                fields.append(f)

    # For Asset and Asset Assignment, include almost all visible, non-read-only fields
    if doctype in ("Asset", "Asset Assignment"):
        all_visible_fields = [
            df.fieldname for df in meta.fields 
            if not df.hidden and not df.read_only 
            and df.fieldtype not in ("Section Break", "Column Break", "Tab Break", "HTML", "Button")
        ]
        for f in all_visible_fields:
            if f not in fields:
                fields.append(f)

    if "name" not in fields and doctype not in ("Lead", "Contacts", "Accounts", "Asset", "Asset Assignment", "Attendance"):
        fields.insert(0, "name")

    export_fields = {doctype: fields}

    # Use standard Exporter to get the template structure
    e = Exporter(
        doctype,
        export_fields=export_fields,
        export_data=False,
        file_type="Excel"
    )
    csv_array = e.get_csv_array_for_export()

    # Create Workbook manually to set formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doctype[:30] # Excel sheet name limit

    header = csv_array[0]
    phone_indices = [i for i, label in enumerate(header) if "(+91-)" in label]

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add Sample Data for Attendance
    if doctype == "Attendance":
        try:
            # Bypass cache completely by using direct SQL query to fetch the freshest record
            latest_attendance = frappe.db.sql("""
                SELECT employee, employee_name, attendance_date, status, in_time, out_time 
                FROM `tabAttendance` 
                ORDER BY modified DESC 
                LIMIT 1
            """, as_dict=True)

            if latest_attendance:
                sample = latest_attendance[0]
                sample_row = []
                for col in header:
                    val = ""
                    if col == "Employee":
                        val = sample.get("employee") or ""
                    elif col == "Employee Name":
                        val = sample.get("employee_name") or ""
                    elif col == "Date":
                        val = str(sample.get("attendance_date")) if sample.get("attendance_date") else ""
                    elif col == "Status":
                        val = sample.get("status") or ""
                    elif col == "In Time":
                        val = str(sample.get("in_time")) if sample.get("in_time") is not None else ""
                    elif col == "Out Time":
                        val = str(sample.get("out_time")) if sample.get("out_time") is not None else ""

                    sample_row.append(val)

                ws.append(sample_row)
            else:
                # Fallback if no records exist yet
                sample_employee = frappe.db.sql("SELECT name, employee_name FROM `tabEmployee` WHERE status='Active' LIMIT 1", as_dict=True)
                if sample_employee:
                    from frappe.utils import nowdate
                    sample = sample_employee[0]
                    sample_row = []
                    for col in header:
                        val = ""
                        if col == "Employee":
                            val = sample.get("name") or ""
                        elif col == "Employee Name":
                            val = sample.get("employee_name") or ""
                        elif col == "Date":
                            val = nowdate()
                        elif col == "Status":
                            val = "Present"
                        elif col == "In Time":
                            val = "09:00:00"
                        elif col == "Out Time":
                            val = "18:00:00"

                        sample_row.append(val)

                    ws.append(sample_row)
        except Exception as e:
            frappe.log_error(f"Error adding sample data for Attendance: {str(e)}")

    # Set column format to Text (@) for phone columns
    # We apply this to the first 100 rows to ensure user input is caught as text
    for col_idx in phone_indices:
        col_letter = get_column_letter(col_idx + 1)
        for r in range(1, 101):
            ws.cell(row=r, column=col_idx + 1).number_format = "@"

    # Save to buffer
    xlsx_file = BytesIO()
    wb.save(xlsx_file)

    # Provide binary response
    provide_binary_file(doctype, "xlsx", xlsx_file.getvalue())


@frappe.whitelist()
def update_import_file(data_import_name, data):
    """
    Save edited preview data back to a CSV file and update the Data Import record.
    """
    import json
    import csv
    import io
    from frappe.utils.file_manager import save_file

    data_import = frappe.get_doc("Data Import", data_import_name)
    data_import.check_permission("write")

    rows = json.loads(data)

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)

    content = output.getvalue()

    # Save as a new file
    filename = f"edited_{data_import_name}.csv"
    file_doc = save_file(
        filename,
        content.encode("utf-8"),
        "Data Import",
        data_import_name,
        is_private=1,
        df="import_file"
    )

    # Update Data Import record
    data_import.import_file = file_doc.file_url
    data_import.save()

    return {"status": "success", "file_url": file_doc.file_url}

@frappe.whitelist()
def get_doctype_fields(doctype):
    meta = frappe.get_meta(doctype)
    fields = []

    for d in meta.fields:
        if d.fieldtype not in ['Section Break', 'Column Break', 'Tab Break', 'HTML', 'Button', 'Table'] and not d.hidden and not d.read_only:
            fields.append({
                "fieldname": d.fieldname,
                "label": d.label,
                "fieldtype": d.fieldtype,
                "options": d.options,
                "hidden": d.hidden
            })

    return {"name": doctype, "fields": fields}

@frappe.whitelist()
def get_hr_dashboard_data():
    """
    Fetch HR dashboard statistics and data.
    """
    data = {}
    today = frappe.utils.today()
    
    # 1. Announcements
    try:
        announcements = frappe.get_all(
            "Announcement",
            filters={"is_active": 1},
            fields=["announcement_name", "announcement", "creation"],
            order_by="creation desc",
            limit=5
        )
        # Map to expected frontend format
        data["announcements"] = [
            {
                "title": a.announcement_name,
                "message": a.announcement,
                "posting_date": str(a.creation.date()) if a.creation else ""
            }
            for a in announcements
        ]
    except Exception:
        data["announcements"] = []

    # 2. Employee Stats
    try:
        data["total_employees"] = frappe.db.count("Employee", {"status": "Active"})
    except Exception:
        data["total_employees"] = 0

    # 3. Pending Leaves (Workflow based)
    try:
        # Assuming workflow states 'Pending' or 'Draft' are pending
        data["pending_leaves"] = frappe.db.count("Leave Application", {
            "workflow_state": "Pending"
        })
    except Exception:
        data["pending_leaves"] = 0

    # 3.1 Pending Requests Count
    try:
        data["pending_request"] = frappe.db.count("Request", {
            "workflow_state": "Pending"
        })
    except Exception:
        data["pending_request"] = 0

    # 5. Today's Leaves
    try:
        # Check Attendance for "On Leave" status or Leave Application for approved leaves today
        data["todays_leaves"] = frappe.db.sql("""
            SELECT e.employee_name, e.name as employee
            FROM `tabLeave Application` la
            JOIN `tabEmployee` e ON la.employee = e.name
            WHERE la.workflow_state = 'Approved'
            AND %s BETWEEN la.from_date AND la.to_date
        """, (today,), as_dict=True)
    except Exception:
        data["todays_leaves"] = []

    # 6. Today's Birthdays
    try:
        # Match day and month
        data["todays_birthdays"] = frappe.db.sql("""
            SELECT employee_name, name as employee
            FROM `tabEmployee`
            WHERE status = 'Active'
            AND DAY(dob) = DAY(%s)
            AND MONTH(dob) = MONTH(%s)
        """, (today, today), as_dict=True)
    except Exception:
        data["todays_birthdays"] = []

    # 8. Pending Leave Applications List
    try:
        data["pending_leaves_list"] = frappe.get_all(
            "Leave Application",
            filters={"workflow_state": "Pending"},
            fields=["name", "employee", "employee_name", "leave_type", "from_date", "to_date", "total_days"],
            order_by="creation desc",
            limit=10
        )
    except Exception:
        data["pending_leaves_list"] = []

    # 9. Pending Requests List
    try:
        data["pending_requests_list"] = frappe.get_all(
            "Request",
            filters={"workflow_state": "Pending"},
            fields=["name", "employee_id", "employee_name", "subject", "creation"],
            order_by="creation desc",
            limit=10
        )
    except Exception:
        data["pending_requests_list"] = []

    return data


@frappe.whitelist()
def convert_estimation_to_invoice(estimation):
    if not estimation:
        frappe.throw("Estimation ID required")

    # Load estimation document
    est = frappe.get_doc("Estimation", estimation)

    # Prevent duplicate conversion
    existing_invoice = frappe.db.get_value(
        "Invoice",
        {"converted_estimation_id": est.name},
        "name"
    )
    if existing_invoice:
        frappe.msgprint(f"Invoice <b>{existing_invoice}</b> already created for this estimation.")
        return existing_invoice

    # Create new invoice
    inv = frappe.new_doc("Invoice")
    inv.flags.ignore_mandatory = True

    # Copy main fields
    fields_to_copy = [
        "customer_name",
        "billing_name",
        "billing_address",
        "phone_number",
        "deal",
        "total_qty",
        "total_amount",
        "overall_discount_type",
        "overall_discount",
        "grand_total",
        "description",
        "terms_and_conditions",
        "bank_account"
    ]

    for f in fields_to_copy:
        inv.set(f, est.get(f))

    # Set client_name (contact ID) from estimation
    inv.client_name = est.client_name

    # Invoice date
    inv.invoice_date = frappe.utils.nowdate()

    # Conversion flags
    inv.converted_from_estimation = 1
    inv.converted_estimation_id = est.name

    # Copy items
    for item in est.get("table_qecz"):
        inv.append("table_qecz", {
            "service": item.service,
            "hsn_code": item.hsn_code,
            "description": item.description,
            "quantity": item.quantity,
            "price": item.price,
            "discount_type": item.discount_type,
            "discount": item.discount,
            "tax_type": item.tax_type,
            "tax_category": item.tax_category,
            "tax_percent": item.tax_percent,
            "tax_amount": item.tax_amount,
            "cgst": item.cgst,
            "sgst": item.sgst,
            "igst": item.igst,
            "sub_total": item.sub_total
        })

    # Save invoice
    inv.insert(ignore_permissions=True, ignore_mandatory=True)

    # SUCCESS MESSAGE
    frappe.msgprint(
        msg=f"Estimation <b>{est.name}</b> successfully converted to Invoice <b>{inv.name}</b>!",
        title="Conversion Complete",
        indicator="green"
    )

    return inv.name


@frappe.whitelist()
def get_sales_dashboard_data(start_date=None, end_date=None):
    """
    Fetch Sales dashboard statistics and data.
    """
    user = frappe.session.user
    # Only filter by owner if the user has a User Permission record restricting their view
    has_user_permission = frappe.db.exists("User Permission", {"user": user, "allow": "User"})
    owner_name = user if has_user_permission else None
    data = {}
    today = frappe.utils.today()
    first_day_month = frappe.utils.get_first_day(today)
    first_day_year = f"{today[:4]}-01-01"

    try:
        # 1. Summary Metrics from Invoices
        invoice_filters = {}
        if start_date and end_date:
            invoice_filters["invoice_date"] = ["between", [start_date, end_date]]
        elif start_date:
            invoice_filters["invoice_date"] = [">=", start_date]
        elif end_date:
            invoice_filters["invoice_date"] = ["<=", end_date]

        if owner_name:
            invoice_filters["owner_name"] = owner_name

        invoices = frappe.get_all("Invoice", filters=invoice_filters, fields=[
            "grand_total", "total_amount", "overall_discount",
            "total_qty", "invoice_date", "balance_amount", "due_date",
            "client_name", "billing_name"
        ])

        data["total_sales"] = sum(frappe.utils.flt(inv.grand_total) for inv in invoices)
        data["total_qty_sold"] = sum(frappe.utils.flt(inv.total_qty) for inv in invoices)
        data["total_orders"] = len(invoices)
        data["aov"] = data["total_sales"] / data["total_orders"] if data["total_orders"] > 0 else 0

        # Gross vs Net
        data["gross_sales"] = sum(frappe.utils.flt(inv.total_amount) for inv in invoices)
        data["net_sales"] = data["total_sales"] # Using grand_total as net sales for now
        data["total_discounts"] = sum(frappe.utils.flt(inv.overall_discount) for inv in invoices)

        # MTD / YTD
        data["mtd_sales"] = sum(frappe.utils.flt(inv.grand_total) for inv in invoices if inv.invoice_date >= frappe.utils.getdate(first_day_month))
        data["ytd_sales"] = sum(frappe.utils.flt(inv.grand_total) for inv in invoices if inv.invoice_date >= frappe.utils.getdate(first_day_year))

        # 2. Pipeline from Deals
        deal_filters = {}
        if start_date and end_date:
            deal_filters["creation"] = ["between", [start_date, f"{end_date} 23:59:59"]]
        elif start_date:
            deal_filters["creation"] = [">=", start_date]
        elif end_date:
            deal_filters["creation"] = ["<=", f"{end_date} 23:59:59"]

        if owner_name:
            deal_filters["owner_name"] = owner_name
            
        deals = frappe.get_all("Deal", filters=deal_filters, fields=["value", "stage"])
        data["pipeline_value"] = sum(frappe.utils.flt(d.value) for d in deals if d.stage not in ["Closed Won", "Closed Lost"])

        # 3. Top Customers
        sql_conds = []
        sql_params = {}
        if start_date and end_date:
            sql_conds.append("DATE(i.invoice_date) BETWEEN %(start_date)s AND %(end_date)s")
            sql_params["start_date"] = start_date
            sql_params["end_date"] = end_date
        elif start_date:
            sql_conds.append("DATE(i.invoice_date) >= %(start_date)s")
            sql_params["start_date"] = start_date
        elif end_date:
            sql_conds.append("DATE(i.invoice_date) <= %(end_date)s")
            sql_params["end_date"] = end_date

        if owner_name:
            sql_conds.append("i.owner_name = %(owner_name)s")
            sql_params["owner_name"] = owner_name

        cond = "WHERE " + " AND ".join(sql_conds) if sql_conds else ""
            
        data["top_customers_by_revenue"] = frappe.db.sql(f"""
            SELECT
                i.client_name,
                i.billing_name,
                COALESCE(NULLIF(CONCAT_WS(' ', c.first_name, c.last_name), ''), c.name) as contact_name,
                a.account_name as account_name,
                SUM(i.grand_total) as revenue,
                COUNT(i.name) as order_count
            FROM `tabInvoice` i
            LEFT JOIN `tabContacts` c ON c.name = i.client_name
            LEFT JOIN `tabAccounts` a ON a.name = i.billing_name
            {cond}
            GROUP BY i.client_name, i.billing_name, c.first_name, c.last_name, a.account_name
            ORDER BY revenue DESC
            LIMIT 5
        """, sql_params, as_dict=True)

        data["most_repeated_customers"] = frappe.db.sql(f"""
            SELECT
                i.client_name,
                i.billing_name,
                COALESCE(NULLIF(CONCAT_WS(' ', c.first_name, c.last_name), ''), c.name) as contact_name,
                a.account_name as account_name,
                COUNT(i.name) as order_count,
                SUM(i.grand_total) as total_spent
            FROM `tabInvoice` i
            LEFT JOIN `tabContacts` c ON c.name = i.client_name
            LEFT JOIN `tabAccounts` a ON a.name = i.billing_name
            {cond}
            GROUP BY i.client_name, i.billing_name, c.first_name, c.last_name, a.account_name
            ORDER BY order_count DESC
            LIMIT 5
        """, sql_params, as_dict=True)

        # 4. Overdue / Pending Orders
        overdue_filters = [
            ["balance_amount", ">", 0],
            ["due_date", "<", today],
            ["due_date", "is", "set"]
        ]
        if start_date and end_date:
            overdue_filters.append(["invoice_date", "between", [start_date, end_date]])
        elif start_date:
            overdue_filters.append(["invoice_date", ">=", start_date])
        elif end_date:
            overdue_filters.append(["invoice_date", "<=", end_date])

        if owner_name:
            overdue_filters.append(["owner_name", "=", owner_name])

        overdue_orders = frappe.get_all("Invoice",
            filters=overdue_filters,
            fields=["name", "billing_name", "due_date", "balance_amount", "grand_total"],
            order_by="due_date asc",
            limit=5
        )

        reordered_overdue = []
        for order in overdue_orders:
            acct_name = frappe.db.get_value("Accounts", order.billing_name, "account_name") or order.billing_name
            # Rebuild dict to ensure account_name appears immediately after billing_name in response
            new_order = {
                "name": order.get("name"),
                "billing_name": order.get("billing_name"),
                "account_name": acct_name,
                "due_date": order.get("due_date"),
                "balance_amount": order.get("balance_amount"),
                "grand_total": order.get("grand_total"),
            }
            reordered_overdue.append(new_order)

        data["overdue_orders"] = reordered_overdue
        
        pending_filters = {"balance_amount": [">", 0]}
        if start_date and end_date:
            pending_filters["invoice_date"] = ["between", [start_date, end_date]]
        elif start_date:
            pending_filters["invoice_date"] = [">=", start_date]
        elif end_date:
            pending_filters["invoice_date"] = ["<=", end_date]

        if owner_name:
            pending_filters["owner_name"] = owner_name
        data["pending_orders_count"] = frappe.db.count("Invoice", pending_filters)

        # 5. Trends
        trend_conds = []
        trend_params = []
        if start_date and end_date:
            trend_conds.append("invoice_date BETWEEN %s AND %s")
            trend_params.extend([start_date, end_date])
        elif start_date:
            trend_conds.append("invoice_date >= %s")
            trend_params.append(start_date)
        elif end_date:
            trend_conds.append("invoice_date <= %s")
            trend_params.append(end_date)
        else:
            # Default to last 12 months
            trend_conds.append("invoice_date >= DATE_SUB(%s, INTERVAL 12 MONTH)")
            trend_params.append(today)

        if owner_name:
            trend_conds.append("owner_name = %s")
            trend_params.append(owner_name)

        trends = frappe.db.sql(f"""
            SELECT
                LEFT(invoice_date, 7) as month,
                SUM(grand_total) as total_sales,
                SUM(overall_discount) as total_discount
            FROM `tabInvoice`
            WHERE {" AND ".join(trend_conds)}
            GROUP BY month
            ORDER BY month ASC
        """, tuple(trend_params), as_dict=True)

        data["sales_trend"] = {
            "categories": [t.month for t in trends],
            "series": [frappe.utils.flt(t.total_sales) for t in trends]
        }

        # 6. Conversion Rate (Estimations to Invoices)
        est_filters = {}
        inv_filters = {"converted_from_estimation": 1}
        if start_date and end_date:
            est_filters["creation"] = ["between", [start_date, f"{end_date} 23:59:59"]]
            inv_filters["creation"] = ["between", [start_date, f"{end_date} 23:59:59"]]
        elif start_date:
            est_filters["creation"] = [">=", start_date]
            inv_filters["creation"] = [">=", start_date]
        elif end_date:
            est_filters["creation"] = ["<=", f"{end_date} 23:59:59"]
            inv_filters["creation"] = ["<=", f"{end_date} 23:59:59"]

        if owner_name:
            est_filters["owner_name"] = owner_name
            inv_filters["owner_name"] = owner_name

        total_estimations = frappe.db.count("Estimation", est_filters)
        converted_estimations = frappe.db.count("Invoice", inv_filters)
        data["conversion_rate"] = (converted_estimations / total_estimations * 100) if total_estimations > 0 else 0

    except Exception as e:
        frappe.log_error(f"Sales Dashboard Error: {str(e)}")
        # Return empty data structure to avoid frontend crashes
        return {
            "total_sales": 0, "total_qty_sold": 0, "total_orders": 0, "aov": 0,
            "gross_sales": 0, "net_sales": 0, "total_discounts": 0,
            "mtd_sales": 0, "ytd_sales": 0, "pipeline_value": 0,
            "top_customers_by_revenue": [], "most_repeated_customers": [],
            "overdue_orders": [], "pending_orders_count": 0,
            "sales_trend": {"categories": [], "series": []},
            "discount_trend": {"categories": [], "series": []},
            "conversion_rate": 0
        }

    return data

@frappe.whitelist(allow_guest=False)
def update_my_password(old_password, new_password):
    """
    Update the current user's password.
    Requires Token Auth: Authorization: token <api_key>:<api_secret>
    """
    user = frappe.session.user

    if not old_password or not new_password:
        return {"status": "failed", "message": "Old password and new password are required."}

    if old_password == new_password:
        return {"status": "failed", "message": "New password must be different from the old password."}

    # Verify old password
    try:
        frappe.utils.password.check_password(user, old_password)
    except frappe.AuthenticationError:
        return {"status": "failed", "message": "Incorrect old password"}

    # Update password and commit to DB
    frappe.utils.password.update_password(user, new_password)
    frappe.db.commit()

    return {
        "status": "success", 
        "message": "Password updated successfully",
        "data": {
            "new-password": new_password,
        }
    }

@frappe.whitelist(allow_guest=True)
def mobile_logout():
    """
    Logout API for Mobile / App.
    Ensures a clean response without extra Frappe fields.
    """
    frappe.local.login_manager.logout()
    
    # Prevent Frappe from appending home_page and full_name to the root of the API response
    frappe.local.response.pop('home_page', None)
    frappe.local.response.pop('full_name', None)
    
    return {"status": "success", "message": "Logged out successfully"}


@frappe.whitelist()
def admin_change_user_password(user_email, new_password):
    """
    Allow administrators to change a user's password.
    Requires System Manager role.
    """
    # Check if current user has permission
    if not frappe.has_permission("User", "write"):
        frappe.throw(_("Not permitted to change user passwords"))
    
    # Verify the target user exists
    if not frappe.db.exists("User", user_email):
        frappe.throw(_("User not found"))
    
    try:
        # Update the password using Frappe's utility
        frappe.utils.password.update_password(user_email, new_password)
        frappe.db.commit()
        
        return {"status": "success", "message": "Password changed successfully"}
    except Exception as e:
        frappe.log_error(f"Error changing password for {user_email}: {str(e)}")
        frappe.throw(_("Failed to change password: {0}").format(str(e)))


@frappe.whitelist(allow_guest=True)
def update_profile_info(first_name, middle_name=None, last_name=None):
    """
    Update the current user's profile information (names).
    """
    user = frappe.session.user

    if user == "Guest":
        return {"status": "failed", "message": "Please login to update profile"}

    try:
        user_doc = frappe.get_doc("User", user)
        user_doc.first_name = first_name
        user_doc.middle_name = middle_name
        user_doc.last_name = last_name
        user_doc.save(ignore_permissions=True)

        return {
            "status": "success",
            "message": "Profile updated successfully",
            "data": {
                "first_name": user_doc.first_name,
                "middle_name": user_doc.middle_name,
                "last_name": user_doc.last_name,
                "full_name": user_doc.full_name
            }
        }
    except Exception as e:
        frappe.log_error(f"Error updating profile: {str(e)}")
        return {"status": "failed", "message": str(e)}


@frappe.whitelist(allow_guest=True)
def upload_profile_image():
    """
    Upload and update profile image for the current user.
    """
    user_email = frappe.session.user

    if user_email == "Guest":
        return {"status": "failed", "message": "Please login to upload profile image"}

    try:
        user_doc = frappe.get_doc("User", user_email)

        # 'file' is the key in FormData
        file = frappe.request.files.get("file")
        if not file:
             return {"status": "failed", "message": "No file uploaded"}

        from frappe.utils.file_manager import save_file

        # Save the file
        fname = file.filename
        content = file.stream.read()

        # Save file and attach to User document
        saved_file = save_file(
            fname,
            content,
            "User",
            user_email,
            decode=False,
            is_private=0,
            df="user_image"
        )

        # Explicitly update user_image just in case save_file df param didn't trigger it (though it should)
        user_doc.user_image = saved_file.file_url
        user_doc.save(ignore_permissions=True)

        return {
            "status": "success",
            "message": "Profile image updated",
            "file_url": saved_file.file_url
        }

    except Exception as e:
        frappe.log_error(f"Error uploading profile image: {str(e)}")
        return {"status": "failed", "message": str(e)}


@frappe.whitelist()
def upload_employee_image():
    """
    Upload and update profile image for the employee linked to the current user.
    Also automatically updates the User record via the Employee on_update hook.
    """
    user_email = frappe.session.user

    if user_email == "Guest":
        return {"status": "failed", "message": "Please login to upload profile image"}

    try:
        # Get Employee ID linked to this user
        employee_id = frappe.db.get_value("Employee", {"user": user_email}, "name")
        if not employee_id:
            return {"status": "failed", "message": "No Employee record found for current user"}

        # 'file' is the key in FormData
        file = frappe.request.files.get("file")
        if not file:
             return {"status": "failed", "message": "No file uploaded"}

        from frappe.utils.file_manager import save_file

        # Save the file and attach to Employee record
        fname = file.filename
        content = file.stream.read()

        saved_file = save_file(
            fname,
            content,
            "Employee",
            employee_id,
            decode=False,
            is_private=0,
            df="profile_picture"
        )

        # Update employee record
        # This will trigger the 'on_update' hook in employee.py which syncs to the User record
        employee_doc = frappe.get_doc("Employee", employee_id)
        employee_doc.profile_picture = saved_file.file_url
        employee_doc.save(ignore_permissions=True)

        return {
            "status": "success",
            "message": "Employee profile image updated and synced to User",
            "file_url": saved_file.file_url
        }

    except Exception as e:
        frappe.log_error(f"Error uploading employee image: {str(e)}")
        return {"status": "failed", "message": str(e)}


@frappe.whitelist()
def get_financial_totals(start_date=None, end_date=None):
    """
    Fetch financial totals for Invoices, Estimations, Purchases, and Expenses.
    Includes total amount, count, and 7-day trend chart data.
    """
    data = {}

    # Get last 7 days for chart
    days = []
    for i in range(6, -1, -1):
        date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
        day_name = frappe.utils.get_datetime(date).strftime('%a')
        days.append(day_name)

    # 1. Invoices
    try:
        cond = ""
        if start_date and end_date:
            cond = f"WHERE DATE(invoice_date) BETWEEN '{start_date}' AND '{end_date}'"
        invoices_total = frappe.db.sql(f"""
            SELECT SUM(grand_total) as total, COUNT(*) as count
            FROM `tabInvoice`
            {cond}
        """, as_dict=True)[0]

        # Chart data for last 7 days (count of invoices created each day)
        invoice_chart = []
        for i in range(6, -1, -1):
            date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
            count = frappe.db.sql("""
                SELECT COUNT(*) as count
                FROM `tabInvoice`
                WHERE DATE(invoice_date) = %s
            """, (date,), as_dict=True)[0]
            invoice_chart.append(count.get('count') or 0)

        data["invoices"] = {
            "total": frappe.utils.flt(invoices_total.get('total') or 0),
            "count": invoices_total.get('count') or 0,
            "chart": invoice_chart
        }
    except Exception as e:
        frappe.log_error(f"Error fetching invoice totals: {str(e)}")
        data["invoices"] = {"total": 0, "count": 0, "chart": [0]*7}

    # 2. Estimations
    try:
        cond = ""
        if start_date and end_date:
            cond = f"WHERE DATE(estimate_date) BETWEEN '{start_date}' AND '{end_date}'"
        estimations_total = frappe.db.sql(f"""
            SELECT SUM(grand_total) as total, COUNT(*) as count
            FROM `tabEstimation`
            {cond}
        """, as_dict=True)[0]

        # Chart data for last 7 days (count of estimations created each day)
        estimation_chart = []
        for i in range(6, -1, -1):
            date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
            count = frappe.db.sql("""
                SELECT COUNT(*) as count
                FROM `tabEstimation`
                WHERE DATE(estimate_date) = %s
            """, (date,), as_dict=True)[0]
            estimation_chart.append(count.get('count') or 0)

        data["estimations"] = {
            "total": frappe.utils.flt(estimations_total.get('total') or 0),
            "count": estimations_total.get('count') or 0,
            "chart": estimation_chart
        }
    except Exception as e:
        frappe.log_error(f"Error fetching estimation totals: {str(e)}")
        data["estimations"] = {"total": 0, "count": 0, "chart": [0]*7}

    # 3. Purchases
    try:
        cond = ""
        if start_date and end_date:
            cond = f"WHERE DATE(purchase_date) BETWEEN '{start_date}' AND '{end_date}'"
        purchases_total = frappe.db.sql(f"""
            SELECT SUM(grand_total) as total, COUNT(*) as count
            FROM `tabPurchase`
            {cond}
        """, as_dict=True)[0]

        # Chart data for last 7 days (count of purchases created each day)
        purchase_chart = []
        for i in range(6, -1, -1):
            date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
            count = frappe.db.sql("""
                SELECT COUNT(*) as count
                FROM `tabPurchase`
                WHERE DATE(purchase_date) = %s
            """, (date,), as_dict=True)[0]
            purchase_chart.append(count.get('count') or 0)

        data["purchases"] = {
            "total": frappe.utils.flt(purchases_total.get('total') or 0),
            "count": purchases_total.get('count') or 0,
            "chart": purchase_chart
        }
    except Exception as e:
        frappe.log_error(f"Error fetching purchase totals: {str(e)}")
        data["purchases"] = {"total": 0, "count": 0, "chart": [0]*7}

    # 4. Expenses
    try:
        cond = ""
        if start_date and end_date:
            cond = f"WHERE DATE(date) BETWEEN '{start_date}' AND '{end_date}'"
        expenses_total = frappe.db.sql(f"""
            SELECT SUM(total) as total, COUNT(*) as count
            FROM `tabExpenses`
            {cond}
        """, as_dict=True)[0]

        # Chart data for last 7 days (count of expenses created each day)
        expense_chart = []
        for i in range(6, -1, -1):
            date = frappe.utils.add_days(frappe.utils.nowdate(), -i)
            count = frappe.db.sql("""
                SELECT COUNT(*) as count
                FROM `tabExpenses`
                WHERE DATE(date) = %s
            """, (date,), as_dict=True)[0]
            expense_chart.append(count.get('count') or 0)

        data["expenses"] = {
            "total": frappe.utils.flt(expenses_total.get('total') or 0),
            "count": expenses_total.get('count') or 0,
            "chart": expense_chart
        }
    except Exception as e:
        frappe.log_error(f"Error fetching expense totals: {str(e)}")
        data["expenses"] = {"total": 0, "count": 0, "chart": [0]*7}

    data["categories"] = days

    return data

@frappe.whitelist()
def get_crm_expense_tracker_stats(start_date=None, end_date=None):
    """
    Get dashboard stats for CRM Expense Tracker.
    Calculates total income, total expense and balance based on the period.
    """
    filters = {}
    if start_date and end_date:
        filters["date_time"] = ["between", [start_date, end_date]]

    stats = {
        "total_income": 0,
        "total_expense": 0,
        "balance": 0
    }

    try:
        data = frappe.get_all("CRM Expense Tracker", filters=filters, fields=["type", "amount"])

        for d in data:
            if d.type == "Income":
                stats["total_income"] += frappe.utils.flt(d.amount)
            elif d.type == "Expense":
                stats["total_expense"] += frappe.utils.flt(d.amount)

        stats["balance"] = stats["total_income"] - stats["total_expense"]
    except Exception as e:
        frappe.log_error(f"Error fetching CRM Expense Tracker stats: {str(e)}")

    return stats

@frappe.whitelist()
def apply_workflow_action(doctype, name, action, comment=None, payment_details=None, update_data=None):
    """
    Apply a workflow action to a document.
    """
    if not frappe.has_permission(doctype, "write", doc=name):
        frappe.throw(_("Not permitted to update this document"), frappe.PermissionError)

    from frappe.model.workflow import apply_workflow, get_transitions
    doc = frappe.get_doc(doctype, name)

    # Debug logging
    try:
        transitions = get_transitions(doc)
        allowed_actions = [t.action for t in transitions]
        frappe.log_error(f"Applying workflow action: {action} on {name}. User: {frappe.session.user}. Roles: {frappe.get_roles()}. Allowed actions: {allowed_actions}", "Workflow Debug")
    except Exception as e:
        frappe.log_error(f"Error in workflow debug logging: {str(e)}", "Workflow Debug Error")

    # Handle optional comment
    if comment:
        doc.add_comment("Workflow", comment)

    # Handle payment details if provided
    if payment_details:
        import json
        if isinstance(payment_details, str):
            pd = json.loads(payment_details)
        else:
            pd = payment_details
        
        if "payment_reference" in pd:
            doc.payment_reference = pd.get("payment_reference")
        if "paid_date" in pd:
            doc.paid_date = pd.get("paid_date")
        if "paid_by" in pd:
            doc.paid_by = pd.get("paid_by")
        
    # Handle optional update_data
    if update_data:
        import json
        if isinstance(update_data, str):
            ud = json.loads(update_data)
        else:
            ud = update_data
        doc.update(ud)
        doc.save(ignore_permissions=True) # Save updates before advancing workflow
        
    apply_workflow(doc, action)
    
    # Reload doc to ensure we get the latest state including paid field
    doc.reload()

    # Push real-time update so Leave Application / Leave Allocation view refreshes instantly
    if doctype == "Leave Application":
        frappe.publish_realtime(
            event="leave_application_updated",
            message={"name": name, "action": action},
        )
    elif doctype == "Leave Allocation":
        frappe.publish_realtime(
            event="leave_allocation_updated",
            message={"name": name, "action": action},
        )
    elif doctype == "Reimbursement Claim":
        frappe.publish_realtime(
            event="reimbursement_claim_updated",
            message={"name": name, "action": action},
        )
    elif doctype == "WFH Attendance":
        frappe.publish_realtime(
            event="wfh_attendance_updated",
            message={"name": name, "action": action},
        )

    return {"status": "success", "message": f"Action {action} applied successfully"}


def _get_attendance_status(hours, p_threshold, h_threshold):
    """Internal helper to categorize status based on hours for Monthly Overview."""
    hours = float(hours or 0)
    p_threshold = float(p_threshold or 5.0)
    h_threshold = float(h_threshold or 3.0)
    
    if hours >= p_threshold:
        return "Present"
    elif hours >= h_threshold:
        return "Half Day"
    else:
        return "Absent"

@frappe.whitelist()
def check_leave_overlap(employee, from_date, to_date, exclude_doc=None):
    """
    Explicitly check for overlapping approved leaves.
    Used by the frontend before applying workflow actions.
    """
    from company.company.api import has_approved_leave
    
    if has_approved_leave(employee, from_date, to_date, exclude_doc=exclude_doc):
        return {
            "overlap": True,
            "message": f"Conflict Detected: Employee {employee} already has an approved leave overlapping this period."
        }
    
    return {"overlap": False}

@frappe.whitelist()
def get_employee_dashboard_data(attendance_range="This Month"):
    """
    Fetch personal dashboard statistics and data for the currently logged-in employee.
    Matches the backend Employee Dashboard layout exactly.
    """
    user = frappe.session.user
    employee_info = frappe.db.get_value("Employee", {"user": user}, ["name", "employee_name"], as_dict=True)
    
    if not employee_info:
        return {}

    employee = employee_info.name
    today = frappe.utils.today()
    
    # Calculate start, end dates and total days in period
    curr_date = frappe.utils.getdate(today)
    month = curr_date.month
    year = curr_date.year
    start_date = frappe.utils.get_first_day(curr_date)
    end_date = today
    total_days_in_period = frappe.utils.get_last_day(curr_date).day # Default: Total days in month
    
    if attendance_range == "Today":
        start_date = today
        total_days_in_period = 1
    elif attendance_range == "This Week":
        start_date = frappe.utils.get_first_day_of_week(curr_date)
        end_date = frappe.utils.add_days(start_date, 6) # End of week
        total_days_in_period = 7
    elif attendance_range == "This Month":
        start_date = frappe.utils.get_first_day(curr_date)
        end_date = frappe.utils.get_last_day(curr_date)
        total_days_in_period = frappe.utils.getdate(end_date).day
    
    # Get probation info
    probation_info = get_employee_probation_info(employee)
    
    data = {
        "employee_name": employee_info.employee_name,
        "employee": employee,
        "attendance_range": attendance_range,
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_days_in_period": total_days_in_period,
        "in_probation": probation_info.get("is_probation", False)
    }

    # 1. Last 7 Days Attendance with Check-in/Out Times
    try:
        seven_days_ago = frappe.utils.add_days(today, -6)  # Include today

        # Get Source Setting
        settings = frappe.get_doc("HRMS Settings")
        source = settings.get("weekly_chart_source") or "Attendance"
        p_threshold = settings.get("present_threshold") or 5.0
        h_threshold = settings.get("half_day_threshold") or 3.0
        
        attendance_map = {}

        if source == "Daily Log":
            # Get Session records for last 7 days
            session_records = frappe.db.sql("""
                SELECT 
                    login_date as date,
                    status,
                    login_time,
                    logout_time,
                    total_work_hours as working_hours
                FROM `tabEmployee Session`
                WHERE employee = %s
                AND login_date >= %s
                AND login_date <= %s
                ORDER BY login_date DESC
            """, (employee, seven_days_ago, today), as_dict=True)

            for record in session_records:
                date_str = frappe.utils.get_date_str(record.date)
                # For Daily Log, we must calculate status dynamically using thresholds
                attendance_map[date_str] = {
                    "status": _get_attendance_status(record.working_hours, p_threshold, h_threshold),
                    "in_time": record.login_time.strftime("%H:%M:%S") if record.login_time else None,
                    "out_time": record.logout_time.strftime("%H:%M:%S") if record.logout_time else None,
                    "working_hours": record.working_hours or 0
                }
        else:
            # Get attendance records for last 7 days
            attendance_records = frappe.db.sql("""
                SELECT 
                    attendance_date as date,
                    status,
                    in_time,
                    out_time,
                    working_hours_decimal as working_hours
                FROM `tabAttendance`
                WHERE employee = %s
                AND attendance_date >= %s
                AND attendance_date <= %s
                ORDER BY attendance_date DESC
            """, (employee, seven_days_ago, today), as_dict=True)
            
            # Helper function to convert timedelta to time string
            def timedelta_to_time_str(td):
                if not td:
                    return None
                total_seconds = int(td.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            for record in attendance_records:
                date_str = frappe.utils.get_date_str(record.date)
                attendance_map[date_str] = {
                    "status": record.status,
                    "in_time": timedelta_to_time_str(record.in_time),
                    "out_time": timedelta_to_time_str(record.out_time),
                    "working_hours": record.working_hours or 0
                }
        
        # Get holiday records for last 7 days using get_all (safer than get_list)
        holiday_records = frappe.get_all(
            "Holidays",
            filters={
                "holiday_date": ["between", [seven_days_ago, today]]
            },
            fields=["holiday_date", "description", "is_working_day"]
        )
        
        holiday_map = {
            frappe.utils.get_date_str(h.holiday_date): {
                "description": h.description,
                "is_working_day": h.is_working_day
            } 
            for h in holiday_records
        }
        
        # Build last 7 days array with all dates (Latest First)
        weekly_attendance = []
        for i in range(0, 7):  # today back to 6 days ago
            date = frappe.utils.add_days(today, -i)
            # Use frappe.utils.get_date_str to ensure consistent YYYY-MM-DD format
            date_str = frappe.utils.get_date_str(date)
            
            # Base data
            day_data = {
                "date": date_str,
                "status": "Not Marked",
                "check_in": None,
                "check_out": None,
                "working_hours": 0,
                "holiday_info": None,
                "holiday_is_working_day": 0
            }
            
            # Add attendance data if exists
            if date_str in attendance_map:
                record = attendance_map[date_str]
                day_data.update({
                    "status": record["status"],
                    "check_in": record["in_time"],
                    "check_out": record["out_time"],
                    "working_hours": record["working_hours"]
                })
            
            # Add holiday info if exists
            if date_str in holiday_map:
                holiday = holiday_map[date_str]
                day_data["holiday_info"] = f"Holiday: {holiday['description']}"
                day_data["holiday_is_working_day"] = holiday["is_working_day"]
                if day_data["status"] == "Not Marked" and not holiday["is_working_day"]:
                    day_data["status"] = "Holiday"
            
            weekly_attendance.append(day_data)
        
        data["weekly_attendance"] = weekly_attendance
        data["weekly_chart_source"] = source
    except Exception as e:
        import traceback
        frappe.log_error(f"Error fetching weekly attendance:\n{traceback.format_exc()}", "Dashboard Error")
        data["weekly_attendance"] = []

    # 2. Leave Allocations by Type (for Leave Status cards)
    try:
        leave_allocations = frappe.db.sql("""
            SELECT
                la.leave_type,
                la.total_leaves_allocated,
                la.total_leaves_taken,
                lt.is_paid
            FROM `tabLeave Allocation` la
            INNER JOIN `tabLeave Type` lt
                ON lt.name = la.leave_type
            WHERE
                la.employee = %s
                AND la.from_date <= %s
                AND la.to_date >= %s
                AND la.docstatus < 2
                AND lt.status = 'Active'
        """, (employee, today, today), as_dict=True)

        paid_allocated = 0
        paid_taken = 0

        unpaid_allocated = 0
        unpaid_taken = 0

        permission_allocated = 0
        permission_taken = 0

        for l in leave_allocations:

            allocated = flt(l.total_leaves_allocated)
            taken = flt(l.total_leaves_taken)

            # Permission (separate)
            if l.leave_type == "Permission":
                permission_allocated += allocated
                permission_taken += taken

            # Paid Leave (sum all paid leave types)
            elif l.is_paid:
                paid_allocated += allocated
                paid_taken += taken

            # Unpaid Leave (sum every non-paid leave except Permission)
            else:
                unpaid_allocated += allocated
                unpaid_taken += taken

        data["leave_allocations"] = [
            {
                "leave_type": "Paid Leave",
                "total_leaves_allocated": paid_allocated,
                "total_leaves_taken": paid_taken,
                "unused_leaves": paid_allocated - paid_taken,
            },
            {
                "leave_type": "Unpaid Leave",
                "total_leaves_allocated": unpaid_allocated,
                "total_leaves_taken": unpaid_taken,
                "unused_leaves": unpaid_allocated - unpaid_taken,
            },
            {
                "leave_type": "Permission",
                "total_leaves_allocated": permission_allocated,
                "total_leaves_taken": permission_taken,
                "unused_leaves": permission_allocated - permission_taken,
            },
        ]

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Error fetching leave allocations")
        data["leave_allocations"] = []


    # -------------------- HOLIDAY DEBUG --------------------
    try:

        # Show all holiday lists in DB
        all_holiday_lists = frappe.db.get_all(
            "Holiday List",
            fields=["name", "month_year", "year"]
        )

        # Try to fetch holiday list using month + year
        holiday_list = frappe.db.get_value(
            "Holiday List",
            {
                "month_year": str(month),
                "year": year
            },
            "name"
        )


        holidays_data = []

        if holiday_list:
            holidays_data = frappe.db.sql("""
                SELECT holiday_date AS date, description, is_working_day, parent
                FROM `tabHolidays`
                WHERE parent = %s
                AND is_working_day = 0
                ORDER BY holiday_date
            """, holiday_list, as_dict=True)


        holiday_dates = {
            str(h.date) for h in holidays_data if int(h.is_working_day or 0) == 0
        }


        data["holidays"] = holidays_data

    except Exception as e:
        holiday_dates = set()
        data["holidays"] = []


    # -------------------- ATTENDANCE DEBUG --------------------
    try:
        total_days = data.get("total_days_in_period", 1)
        data["hide_missing"] = True if source == "Daily Log" else False

        # Show raw attendance rows
        raw_attendance = frappe.db.sql("""
            SELECT attendance_date, status
            FROM `tabAttendance`
            WHERE employee = %s
            AND attendance_date BETWEEN %s AND %s
            ORDER BY attendance_date
        """, (employee, start_date, end_date), as_dict=True)


        breakdown = {
            "present": 0,
            "absent": 0,
            "half_day": 0,
            "on_leave": 0,
            "missing": 0,
            "holiday": len(holiday_dates),
            "total_days": 0,
            "calendar_total": total_days
        }

        # Dynamic Breakdown Calculation based on source
        if source == "Daily Log":
            attendance_records = frappe.db.sql("""
                SELECT total_work_hours as working_hours
                FROM `tabEmployee Session`
                WHERE employee = %s
                AND login_date BETWEEN %s AND %s
            """, (employee, start_date, end_date), as_dict=True)
            
            for record in attendance_records:
                status = _get_attendance_status(record.working_hours, p_threshold, h_threshold)
                status_key = status.lower().replace(" ", "_")
                if status_key in breakdown:
                    breakdown[status_key] += 1
        else:
            # For Attendance, we use the pre-calculated status via SQL grouping (original logic)
            attendance_breakdown = frappe.db.sql("""
                SELECT 
                    status,
                    COUNT(*) as count
                FROM `tabAttendance`
                WHERE employee = %s
                AND attendance_date BETWEEN %s AND %s
                GROUP BY status
            """, (employee, start_date, end_date), as_dict=True)

            for record in attendance_breakdown:
                status_val = record.get("status")
                if status_val:
                    status_key = status_val.lower().replace(" ", "_").replace("-", "_")
                    if status_key == "leave":
                        status_key = "on_leave"
                    if status_key in breakdown:
                        breakdown[status_key] = int(record.get("count") or 0)

        workingDays = total_days - breakdown["holiday"]
        breakdown["total_days"] = workingDays

        # Calculate "Till Today" metrics for the percentage
        today_date = frappe.utils.getdate(today)
        start_date_obj = frappe.utils.getdate(start_date)
        
        # Calendar days elapsed so far in the period
        days_elapsed = frappe.utils.date_diff(today_date, start_date_obj) + 1
        
        # Holidays that have already occurred in the period
        holidays_elapsed = len([
            d for d in holiday_dates 
            if frappe.utils.getdate(d) <= today_date
        ])

        # Weighted calculation Till Today
        # Numerator includes user-requested components
        present_weighted = (
            breakdown["present"] + 
            breakdown["on_leave"] + 
            (breakdown["half_day"] * 0.5) +
            breakdown["missing"] + 
            holidays_elapsed  # Use elapsed holidays in the numerator
        )

        # Denominator is total calendar days elapsed
        breakdown["attendance_percentage"] = (
            round((present_weighted / days_elapsed) * 100)
            if days_elapsed > 0 else 0
        )

        data["monthly_attendance_breakdown"] = breakdown

    except Exception as e:

        data["monthly_attendance_breakdown"] = {
            "present": 0,
            "absent": 0,
            "half_day": 0,
            "on_leave": 0,
            "missing": 0,
            "total_days": 0
        }

    # 4. Missing Timesheets (Current Month-to-Date - All Working Days)  
    try:
        today_date = frappe.utils.getdate(today)
        month_start_date = frappe.utils.get_first_day(today_date)
        curr_ptr = month_start_date
        potential_working_days = []
        
        # Use the holiday dates already calculated from the Calendar logic
        month_holidays = list(holiday_dates) if 'holiday_dates' in locals() else []

        # Get leave dates for this month
        leave_dates = []
        try:
            leaves = frappe.get_all("Leave Application",
                filters={
                    "employee": employee,
                    "status": "Approved",
                    "from_date": ["<=", today_date],
                    "to_date": [">=", month_start_date],
                    "docstatus": 1
                },
                fields=["from_date", "to_date"]
            )
            for l in leaves:
                d_ptr = max(l.from_date, month_start_date)
                d_end = min(l.to_date, today_date)
                while d_ptr <= d_end:
                    leave_dates.append(str(d_ptr))
                    d_ptr = frappe.utils.add_days(d_ptr, 1)
        except Exception:
            pass

        # Also get dates marked as 'On Leave' in attendance
        try:
            attendance_leaves = frappe.get_all("Attendance",
                filters={
                    "employee": employee,
                    "status": "On Leave",
                    "attendance_date": ["between", [month_start_date, today_date]]
                },
                pluck="attendance_date"
            )
            leave_dates.extend([str(d) for d in attendance_leaves])
        except Exception:
            pass
        
        leave_dates_set = set(leave_dates)

        # Generate all past/current days in month
        while curr_ptr <= today_date:
            # Skip Sundays (frappe.utils.get_weekday returns 'Monday', 'Tuesday', etc.)
            if frappe.utils.get_weekday(curr_ptr) != 'Sunday':
                date_str = str(curr_ptr)
                if date_str not in month_holidays and date_str not in leave_dates_set:
                    potential_working_days.append(date_str)
            # Safely increment as date object
            curr_ptr = frappe.utils.add_days(curr_ptr, 1)
            if isinstance(curr_ptr, str):
                curr_ptr = frappe.utils.getdate(curr_ptr)

        # Get dates with timesheets
        timesheet_dates = frappe.db.sql("""
            SELECT DISTINCT timesheet_date
            FROM `tabTimesheet`
            WHERE employee = %s
            AND timesheet_date >= %s
            AND timesheet_date <= %s
            AND docstatus < 2
        """, (employee, month_start_date, today_date), as_list=True)
        
        timesheet_dates_set = {str(date[0]) for date in timesheet_dates}
        
        # Find missing dates from working days
        missing_dates = sorted([d for d in potential_working_days if d not in timesheet_dates_set])
        data["missing_timesheets"] = [{"date": date} for date in missing_dates]
        data["leave_dates"] = list(leave_dates_set)
        
    except Exception as e:
        import traceback
        frappe.log_error(traceback.format_exc(), "Missing Timesheets Error")
        data["missing_timesheets"] = []

    # 5. Recent Leave Applications
    try:
        data["recent_leaves"] = frappe.get_all("Leave Application",
            filters={"employee": employee},
            fields=["name", "leave_type", "from_date", "to_date", "workflow_state", "total_leave_days"],
            order_by="creation desc",
            limit=5
        )
    except Exception:
        data["recent_leaves"] = []

    # 6. Global Data (Announcements, Birthdays, Holidays, Today's Leaves)
    hr_data = get_hr_dashboard_data()
    data["announcements"] = hr_data.get("announcements", [])
    data["todays_birthdays"] = hr_data.get("todays_birthdays", [])
    data["todays_leaves"] = hr_data.get("todays_leaves", [])

    # 7. Attendance for Calendar (Last 6 months + Current range)
    try:
        # Fetch a broader range for the calendar to show "all data present"
        calendar_start = frappe.utils.add_months(start_date, -6)
        
        full_attendance = frappe.get_all("Attendance",
            filters={
                "employee": employee,
                "attendance_date": ["between", [calendar_start, end_date]]
            },
            fields=["attendance_date", "status", "in_time", "out_time", "working_hours_decimal as working_hours"],
            order_by="attendance_date asc"
        )
        
        # Helper to convert timedelta to time str
        def td_to_str(td):
            if not td: return None
            total_seconds = int(td.total_seconds())
            return f"{total_seconds // 3600:02d}:{(total_seconds % 3600) // 60:02d}:{(total_seconds % 60):02d}"

        # Create attendance map
        attendance_map = {}
        for att in full_attendance:
            attendance_map[str(att.attendance_date)] = {
                "status": att.status,
                "check_in": td_to_str(att.in_time),
                "check_out": td_to_str(att.out_time),
                "working_hours": att.working_hours or 0
            }

        # Build full month timeline
        attendance_events = []
        month_start = frappe.utils.get_first_day(today_date)
        month_end = frappe.utils.get_last_day(today_date)
        
        # Get holidays for this period correctly
        holiday_list = None
        company = frappe.db.get_value("Employee", employee, "company")
        if company:
            holiday_list = frappe.db.get_value("Company", company, "default_holiday_list")
        if not holiday_list:
            holiday_list = frappe.db.get_value("Holiday List", {}, "name")
        
        holiday_map = {}
        if holiday_list:
            h_records = frappe.db.sql("""
                SELECT holiday_date as date, description, is_working_day
                FROM `tabHolidays`
                WHERE parent = %s
                AND holiday_date BETWEEN %s AND %s
            """, (holiday_list, month_start, month_end), as_dict=True)
            holiday_map = {str(h.date): h for h in h_records}

        curr_ptr = month_start
        while curr_ptr <= month_end:
            d_str = str(curr_ptr)
            day_record = {
                "date": d_str,
                "status": "Not Marked",
                "check_in": None,
                "check_out": None,
                "working_hours": 0,
                "holiday_info": None,
                "holiday_is_working_day": 0
            }
            
            if d_str in attendance_map:
                day_record.update(attendance_map[d_str])
            
            if d_str in holiday_map:
                h = holiday_map[d_str]
                day_record["holiday_info"] = h["description"]
                day_record["holiday_is_working_day"] = h["is_working_day"]
                if day_record["status"] == "Not Marked" and not h["is_working_day"]:
                    day_record["status"] = "Holiday"
            
            attendance_events.append(day_record)
            curr_ptr = frappe.utils.add_days(curr_ptr, 1)
            if isinstance(curr_ptr, str):
                curr_ptr = frappe.utils.getdate(curr_ptr)

        data["monthly_attendance_list"] = attendance_events
    except Exception as e:
        data["monthly_attendance_list"] = []

    return data

@frappe.whitelist()
def get_personality_dashboard_data(employee=None):
    """
    Fetch personality dashboard data for the specified employee or the logged-in employee.
    Sends the current total score, the last 5 evaluation traits,
    and a tallied list of improvement suggestions.
    """
    if employee:
        employee_doc = frappe.db.get_value("Employee", employee, ["name", "evaluation_score", "evaluation_status"], as_dict=True)
    else:
        user = frappe.session.user
        employee_doc = frappe.db.get_value("Employee", {"user": user}, ["name", "evaluation_score", "evaluation_status"], as_dict=True)

    if not employee_doc:
        return {
            "totalScore": 100,
            "status": "Excellent",
            "traits": [],
            "performance_breakdown": [],
            "recent_evaluations": []
        }

    employee = employee_doc

    # Fetch all evaluations for calculating dynamic trait scores and recent list
    all_evals = frappe.db.sql("""
        SELECT name, trait, score_change, how_to_improve, creation
        FROM `tabEmployee Evaluation`
        WHERE employee = %s AND docstatus = 1
        ORDER BY creation DESC
    """, (employee.name,), as_dict=True)

    if not all_evals:
        # Initialize default breakdown if there are no evaluations
        db_traits = frappe.db.get_all("Evaluation Trait", fields=["name", "trait_name"])
        performance_breakdown = [{
            "trait": t.trait_name or t.name,
            "score": 100,
            "impact": 0
        } for t in db_traits]
        return {
            "totalScore": 100,
            "status": "Excellent",
            "lastUpdated": None,
            "traits": [],
            "performance_breakdown": performance_breakdown,
            "recent_evaluations": []
        }

    # 1. Prepare traits list (compatibility/fallback)
    trait_scores = []
    for ev in all_evals[:5]:
        trait_name = frappe.db.get_value("Evaluation Trait", ev.trait, "trait_name")
        trait_scores.append({
            "trait": trait_name or ev.trait,
            "score": ev.score_change or 0
        })

    # 2. Calculate dynamic performance breakdown (cumulative scores for all traits)
    db_traits = frappe.db.get_all("Evaluation Trait", fields=["name", "trait_name"])
    trait_data = {}
    for t in db_traits:
        trait_data[t.name] = {
            "trait": t.trait_name or t.name,
            "score": 100,
            "impact": 0,
            "latest_seen": False
        }

    for ev in all_evals:
        t_id = ev.trait
        if t_id in trait_data:
            # Accumulate score change
            trait_data[t_id]["score"] += (ev.score_change or 0)
            
            # The first evaluation we see in DESC order is the latest, so it sets the current impact
            if not trait_data[t_id]["latest_seen"]:
                trait_data[t_id]["impact"] = ev.score_change or 0
                trait_data[t_id]["latest_seen"] = True

    # Clamp scores between 0 and 100
    for t_id in trait_data:
        trait_data[t_id]["score"] = max(0, min(100, trait_data[t_id]["score"]))
        # Remove helper flag
        if "latest_seen" in trait_data[t_id]:
            del trait_data[t_id]["latest_seen"]

    performance_breakdown = list(trait_data.values())

    # 3. Compile Recent Evaluations with detail fields
    recent_evaluations = []
    for ev in all_evals[:5]:
        trait_name = frappe.db.get_value("Evaluation Trait", ev.trait, "trait_name") or ev.trait
        recent_evaluations.append({
            "name": ev.name,
            "trait": trait_name,
            "score_change": ev.score_change or 0,
            "creation": str(ev.creation),
            "remarks": ev.how_to_improve or ""
        })

    # 4. Calculate tallies for how_to_improve (Based on last 20 evaluations)
    trait_tally = {}
    for ev in all_evals[:20]:
        t_id = ev.trait
        score = ev.score_change or 0
        if t_id not in trait_tally:
            trait_tally[t_id] = {
                "score_sum": 0,
                "latest_advice": None,
                "date": ev.creation
            }
        
        trait_tally[t_id]["score_sum"] += score
        
        if not trait_tally[t_id]["latest_advice"] and score < 0 and ev.how_to_improve:
            trait_tally[t_id]["latest_advice"] = ev.how_to_improve
            trait_tally[t_id]["date"] = ev.creation

    how_to_improve_list = []
    from frappe.utils import formatdate
    for t_id, counts in trait_tally.items():
        if counts["score_sum"] < 0:
            advice = counts["latest_advice"]
            trait_name = frappe.db.get_value("Evaluation Trait", t_id, "trait_name") or t_id
            
            if not advice:
                advice = frappe.db.get_value("Evaluation Trait", t_id, "how_to_improve")
            
            if advice:
                date_str = formatdate(counts["date"], "dd-mm-yyyy") if counts["date"] else ""
                formatted_advice = f"{advice} - {trait_name} ({date_str})" if date_str else f"{advice} - {trait_name}"
                
                if formatted_advice not in how_to_improve_list:
                    how_to_improve_list.append(formatted_advice)

    last_eval_time = all_evals[0].creation
    
    calculated_total = employee.evaluation_score if employee.evaluation_score is not None else 100
    calculated_total = max(0, min(100, calculated_total))

    status = employee.evaluation_status or "Excellent"
    if not status or status == "":
        if calculated_total >= 90: status = "Excellent"
        elif calculated_total >= 75: status = "Good"
        elif calculated_total >= 60: status = "Average"
        else: status = "Needs Improvement"

    return {
        "totalScore": calculated_total,
        "status": status,
        "howToImprove": how_to_improve_list[:5],
        "lastUpdated": str(last_eval_time),
        "traits": trait_scores,
        "performance_breakdown": performance_breakdown,
        "recent_evaluations": recent_evaluations
    }


@frappe.whitelist()
def get_weekly_present_absent_data(filter_type=None, from_date=None, to_date=None):
    """
    Returns daily present and absent counts for a given filter range (Last 7 Days, This Month, Last Month, Custom).
    Source (Attendance or Daily Log) is determined by HRMS Settings.
    """
    from datetime import timedelta
    from frappe.utils import getdate, nowdate, add_months, get_first_day, get_last_day
    
    today_dt = getdate(nowdate())

    # 1. Determine Date Range
    if filter_type == "This Month":
        start_date = getdate(get_first_day(today_dt))
        end_date = today_dt
    elif filter_type == "Last Month":
        last_month_dt = add_months(today_dt, -1)
        start_date = getdate(get_first_day(last_month_dt))
        end_date = getdate(get_last_day(last_month_dt))
    elif filter_type == "Custom" and from_date and to_date:
        start_date = getdate(from_date)
        end_date = getdate(to_date)
    else: # Default: Last 7 Days
        start_date = today_dt - timedelta(days=6)
        end_date = today_dt

    # 3. Get Settings, Baseline and Holiday dates for the range
    settings = frappe.get_doc("HRMS Settings")
    source = settings.get("weekly_chart_source") or "Attendance"
    total_active_employees = frappe.db.count("Employee", {"status": "Active"})

    all_holidays = set()
    # Robust holiday fetching logic
    try:
        h_records = frappe.db.sql("""
            SELECT DISTINCT holiday_date as date
            FROM `tabHolidays`
            WHERE holiday_date BETWEEN %s AND %s
            AND is_working_day = 0
        """, (start_date, end_date), as_dict=True)
        all_holidays = {str(h.date) for h in h_records}
    except Exception:
        all_holidays = set()

    # 4. Optimized Bulk Fetch
    present_counts_map = {}
    
    # We only fetch if start_date exists and range is reasonable
    if start_date <= end_date:
        if source == "Attendance":
            # For Attendance source, count Present and Half Day statuses
            attendance_data = frappe.db.sql("""
                SELECT attendance_date, COUNT(*) as count
                FROM `tabAttendance`
                WHERE attendance_date BETWEEN %s AND %s
                AND status IN ('Present', 'Half Day')
                GROUP BY attendance_date
            """, (start_date, end_date), as_dict=True)
            present_counts_map = {str(d.attendance_date): d.count for d in attendance_data}
        else: # Daily Log source
            # For Daily Log source, count unique employees per day
            daily_log_data = frappe.db.sql("""
                SELECT login_date, COUNT(DISTINCT employee) as count
                FROM `tabEmployee Session`
                WHERE login_date BETWEEN %s AND %s
                GROUP BY login_date
            """, (start_date, end_date), as_dict=True)
            present_counts_map = {str(d.login_date): d.count for d in daily_log_data}

    # 5. Generate Result Array
    result = []
    num_days = (end_date - start_date).days + 1
    
    # Safety limit to prevent massive ranges from crashing frontend (e.g. 5 years)
    if num_days > 400:
        num_days = 400
        end_date = start_date + timedelta(days=399)

    for i in range(num_days):
        curr_date = start_date + timedelta(days=i)
        curr_date_str = curr_date.strftime('%Y-%m-%d')
        day_name = curr_date.strftime('%a')
        
        # Skip holidays and non-working days
        if curr_date_str in all_holidays:
            continue

        present_count = present_counts_map.get(curr_date_str, 0)
        absent_count = max(0, total_active_employees - present_count)
        
        result.append({
            "date": curr_date_str,
            "day": day_name,
            "present": present_count,
            "absent": absent_count
        })
        
    return result

@frappe.whitelist()
def get_livekit_token(room_name):
    """Generates an Access Token for a LiveKit room."""
    user = frappe.session.user
    
    # Read credentials from HRMS Settings
    try:
        settings = frappe.get_doc("HRMS Settings")
        url = settings.livekit_url
        api_key = settings.livekit_api_key
        api_secret = settings.get_password("livekit_api_secret")
    except Exception:
        frappe.throw("Failed to read LiveKit settings from HRMS Settings")
    
    if not api_key or not api_secret:
        frappe.throw("LiveKit API Key or Secret not configured in HRMS Settings")

    # Define grants (explicitly including publish/subscribe)
    grant = api.VideoGrants(
        room_join=True,
        room=room_name,
        can_publish=True,
        can_subscribe=True,
        can_publish_data=True
    )
    
    # Create clean identity (remove spaces/special chars if any)
    safe_identity = "".join(c for c in user if c.isalnum() or c in "._-")
    
    # Create AccessToken
    token = api.AccessToken(api_key, api_secret) \
        .with_grants(grant) \
        .with_identity(safe_identity) \
        .with_name(frappe.db.get_value("User", user, "full_name") or user)
        
    return {
        "token": token.to_jwt(),
        "server_url": url
    }


@frappe.whitelist()
def get_open_jobs(search=None, filters=None, limit_start=0, limit_page_length=10, order_by="posted_on desc"):
    """Fetch active job openings for employees to refer candidates."""
    frappe_filters = {"status": "Open"}
    or_filters = []
    
    if search:
        or_filters = [
            ["job_title", "like", f"%{search}%"],
            ["designation", "like", f"%{search}%"],
            ["name", "like", f"%{search}%"]
        ]
        
    if filters:
        import json
        if isinstance(filters, str):
            filters = json.loads(filters)
        if filters.get("location") and filters["location"] != "all":
            frappe_filters["location"] = filters["location"]
            
    return frappe.get_all(
        "Job Opening",
        filters=frappe_filters,
        or_filters=or_filters,
        fields=[
            "name", "job_title", "designation", "experience", "location", 
            "small_description", "description", "posted_on", "closes_on", 
            "status", "shift", "lower_range", "upper_range", "salary_per", "skills_required"
        ],
        limit_start=limit_start,
        limit_page_length=limit_page_length,
        order_by=order_by
    )



@frappe.whitelist()
def get_my_referrals(search=None, filters=None, limit_start=0, limit_page_length=10, order_by="creation desc"):
    """Fetch referrals submitted by the currently logged-in employee."""
    user = frappe.session.user
    employee = frappe.db.get_value("Employee", {"user": user}, "name")
    
    if not employee:
        return []
        
    frappe_filters = {"referrer": employee}
    or_filters = []
    
    if search:
        or_filters = [
            ["candidate_name", "like", f"%{search}%"],
            ["candidate_email", "like", f"%{search}%"],
            ["job_opening", "like", f"%{search}%"]
        ]
        
    if filters:
        import json
        if isinstance(filters, str):
            filters = json.loads(filters)
        if filters.get("status") and filters["status"] != "all":
            frappe_filters["status"] = filters["status"]
        if filters.get("job_opening") and filters["job_opening"] != "all":
            frappe_filters["job_opening"] = filters["job_opening"]

    return frappe.get_all(
        "Employee Referral",
        filters=frappe_filters,
        or_filters=or_filters,
        fields=["name", "candidate_name", "candidate_email", "job_opening", "status", "creation", "job_applicant"],
        order_by=order_by,
        limit_start=limit_start,
        limit_page_length=limit_page_length
    )


@frappe.whitelist()
def submit_referral(candidate_name, candidate_email, job_opening, resume, candidate_phone=None, relationship=None, notes=None):
    """Submit a new candidate referral."""
    user = frappe.session.user
    employee = frappe.db.get_value("Employee", {"user": user}, "name")
    
    if not employee:
        frappe.throw(_("Employee record not found for the current user."))
        
    doc = frappe.get_doc({
        "doctype": "Employee Referral",
        "referrer": employee,
        "job_opening": job_opening,
        "candidate_name": candidate_name,
        "candidate_email": candidate_email,
        "candidate_phone": candidate_phone,
        "resume": resume,
        "relationship": relationship,
        "notes": notes,
        "status": "Pending"
    })
    
    doc.insert(ignore_permissions=True)
    return doc.as_dict()

@frappe.whitelist()
def handle_create_job_applicant(referral_name):
    """API wrapper to call create_job_applicant method on Employee Referral doc."""
    doc = frappe.get_doc("Employee Referral", referral_name)
    return doc.create_job_applicant()

# --- ASSET REQUEST APIs ---

@frappe.whitelist()
def submit_asset_request(request_type, priority, purpose, asset_category=None, asset=None, asset_name=None, asset_tag=None, employee=None, return_attachment=None):
    # Check if current user is HR or System Manager
    current_user = frappe.session.user
    user_roles = frappe.get_roles(current_user)
    is_hr = any(role in user_roles for role in ["HR Manager", "HR", "System Manager", "Administrator"])

    target_employee = None

    if is_hr and employee:
        # Validate that the provided employee exists
        if not frappe.db.exists("Employee", employee):
            frappe.throw(_("Employee {0} not found").format(employee))
        target_employee = employee
    else:
        # Default to the employee linked to the session user
        target_employee = frappe.db.get_value("Employee", {"user": current_user}, "name")

    if not target_employee:
        frappe.throw(_("Employee profile not found for the current user. Please select an employee."))

    doc = frappe.get_doc({
        "doctype": "Asset Request",
        "employee": target_employee,
        "request_type": request_type,
        "priority": priority,
        "purpose": purpose,
        "asset_category": asset_category,
        "asset": asset,
        "asset_name": asset_name,
        "asset_tag": asset_tag,
        "return_attachment": return_attachment
    })
    doc.insert(ignore_permissions=True)
    
    # Automatically promote to Pending Approval so it appears in HR dashboard
    doc.workflow_state = "Pending Approval"
    doc.status = "Pending Approval"
    doc.save(ignore_permissions=True)
    
    frappe.publish_realtime(
        event="asset_request_updated",
        message={"name": doc.name, "workflow_state": doc.workflow_state},
    )

    frappe.db.commit()
    return {"message": "Asset Request submitted successfully.", "name": doc.name}

@frappe.whitelist()
def get_my_asset_requests(page=1, limit=10, request_type=None, status=None, sort_by="modified desc", category=None, asset=None, priority=None, start_date=None, end_date=None):
    employee = frappe.db.get_value("Employee", {"user": frappe.session.user}, "name")
    if not employee:
        return {"data": [], "total": 0}
        
    filters = [["Asset Request", "employee", "=", employee]]
    if request_type and request_type != 'all':
        filters.append(["Asset Request", "request_type", "=", request_type])
    if status and status != 'all':
        filters.append(["Asset Request", "status", "=", status])
    if category:
        filters.append(["Asset Request", "asset_category", "=", category])
    if asset:
        filters.append(["Asset Request", "asset", "like", f"%{asset}%"])
    if priority:
        filters.append(["Asset Request", "priority", "=", priority])
    if start_date:
        filters.append(["Asset Request", "modified", ">=", start_date])
    if end_date:
        filters.append(["Asset Request", "modified", "<=", end_date + " 23:59:59"])
        
    start = (int(page) - 1) * int(limit)
    total = frappe.db.count("Asset Request", filters)
    
    requests = frappe.get_all(
        "Asset Request",
        filters=filters,
        fields=["name", "request_type", "asset_category", "asset_name", "asset", "status", "priority", "creation", "modified", "purpose"],
        order_by=sort_by,
        limit_start=start,
        limit_page_length=int(limit)
    )

    # Fetch and populate asset names for Return Requests if empty
    asset_ids = [r.asset for r in requests if r.asset and not r.asset_name]
    if asset_ids:
        asset_map = {a.name: a.asset_name for a in frappe.get_all("Asset", filters={"name": ["in", asset_ids]}, fields=["name", "asset_name"])}
        for r in requests:
            if r.asset and not r.asset_name:
                r.asset_name = asset_map.get(r.asset)

    return {"data": requests, "total": total}

@frappe.whitelist()
def get_pending_asset_requests(page=1, limit=10, request_type=None, status=None, sort_by="modified desc", category=None, asset=None, priority=None, start_date=None, end_date=None, unread_only=None):
    filters = []
    
    # Check unread filter if requested
    if unread_only in ('true', '1', True):
        hr_roles = ["HR", "HR Manager", "System Manager", "Administrator"]
        user_roles = frappe.get_roles(frappe.session.user)
        is_hr = any(role in user_roles for role in hr_roles)
        if is_hr:
            unread_names = frappe.get_all(
                "HR Read Tracker",
                filters={"reference_doctype": "Asset Request", "read_by": frappe.session.user, "is_read": 0},
                pluck="reference_name"
            )
            filters.append(["Asset Request", "name", "in", unread_names if unread_names else [""]])

    if request_type and request_type != 'all':
        filters.append(["Asset Request", "request_type", "=", request_type])
    if status and status != 'all':
        filters.append(["Asset Request", "status", "=", status])
    if category:
        filters.append(["Asset Request", "asset_category", "=", category])
    if asset:
        filters.append(["Asset Request", "asset", "like", f"%{asset}%"])
    if priority:
        filters.append(["Asset Request", "priority", "=", priority])
    if start_date:
        filters.append(["Asset Request", "modified", ">=", start_date])
    if end_date:
        filters.append(["Asset Request", "modified", "<=", end_date + " 23:59:59"])
        
    start = (int(page) - 1) * int(limit)
    total = frappe.db.count("Asset Request", filters)
    
    requests = frappe.get_all(
        "Asset Request",
        filters=filters,
        fields=["name", "employee", "employee_name", "request_type", "asset_category", "asset_name", "asset", "status", "priority", "creation", "modified", "purpose"],
        order_by=sort_by,
        limit_start=start,
        limit_page_length=int(limit)
    )

    # Fetch and populate asset names for Return Requests if empty
    asset_ids = [r.asset for r in requests if r.asset and not r.asset_name]
    if asset_ids:
        asset_map = {a.name: a.asset_name for a in frappe.get_all("Asset", filters={"name": ["in", asset_ids]}, fields=["name", "asset_name"])}
        for r in requests:
            if r.asset and not r.asset_name:
                r.asset_name = asset_map.get(r.asset)

    return {"data": requests, "total": total}

@frappe.whitelist()
def approve_declaration(request_name, hr_remarks=None, asset_name=None, asset_tag=None, asset_category=None, purchase_date=None, purchase_cost=None):
    from frappe.utils import nowdate
    
    # 1. Fetch the request
    req = frappe.get_doc("Asset Request", request_name)
    if req.workflow_state == "Completed":
        return {"message": "Request is already completed."}
        
    # 2. Create the Asset
    asset = frappe.get_doc({
        "doctype": "Asset",
        "asset_name": asset_name,
        "asset_tag": asset_tag,
        "category": asset_category,
        "purchase_date": purchase_date,
        "purchase_cost": purchase_cost,
        "current_status": "Assigned"
    })
    asset.insert(ignore_permissions=True)
    
    # 3. Create the Assignment
    assignment = frappe.get_doc({
        "doctype": "Asset Assignment",
        "asset": asset.name,
        "assigned_to": req.employee,
        "assigned_on": nowdate(),
        "remarks": hr_remarks or "Generated from Asset Declaration"
    })
    assignment.insert(ignore_permissions=True)
    
    # 4. Push Request to Completed directly to bypass default on_update auto-creation
    req.db_set("hr_remarks", hr_remarks)
    req.db_set("asset", asset.name)
    req.db_set("assigned_asset", asset.name)
    req.db_set("status", "Completed")
    req.db_set("workflow_state", "Completed")
    
    frappe.publish_realtime(
        event="asset_request_updated",
        message={"name": req.name, "workflow_state": "Completed"},
    )
    
    frappe.db.commit()
    return {"message": "Declaration approved and Asset registered successfully."}

@frappe.whitelist()
def get_available_assets_list():
    """
    Returns only Assets that reflect truly available status.
    Must have current_status="Available" AND no active Asset Assignment.
    """
    # 1. Get all assets marked as Available
    available_assets = frappe.get_all(
        "Asset",
        filters={"current_status": "Available"},
        fields=["name", "asset_name", "asset_tag", "category"]
    )
    
    # 2. Get names of assets currently assigned (no returned_on date)
    assigned_asset_names = frappe.get_all(
        "Asset Assignment",
        filters={"returned_on": ["is", "not set"]},
        pluck="asset"
    )
    
    # 3. Filter out any that might be assigned but still marked Available (safeguard)
    return [a for a in available_assets if a.name not in assigned_asset_names]

@frappe.whitelist()
def get_my_assigned_assets(employee=None):
    """
    Returns assets currently assigned to the given employee (where returned_on is null).
    """
    if not employee:
        employee = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")
        
    if not employee:
        return []

    # Get active assignments for the employee
    active_assignments = frappe.get_all(
        "Asset Assignment",
        filters={
            "assigned_to": employee,
            "returned_on": ["is", "not set"]
        },
        fields=["asset", "asset_name"]
    )
    
    return active_assignments


@frappe.whitelist()
def get_hr_task_stats(project=None, department=None, from_date=None, to_date=None, employee_id=None):
    """
    Fetch task statistics for the HR Dashboard with optional filtering.
    Returns counts for total, open, in-progress, completed, and on-hold tasks.
    """
    if not frappe.has_permission("Task Manager", "read"):
        return {}

    filters = {}
    if project and project != 'All':
        filters["project"] = project
    if department and department != 'All':
        filters["department"] = department
    
    if from_date and to_date:
        filters["due_date"] = ["between", [from_date, to_date]]

    tasks_list = None
    if employee_id:
        tasks_with_employee = frappe.get_all("Task Manager Assignee", filters={"employee": employee_id}, pluck="parent")
        if tasks_with_employee:
            filters["name"] = ["in", tasks_with_employee]
            tasks_list = tasks_with_employee
        else:
            return {
                "total": 0, "open": 0, "reopen": 0, "in_progress": 0, "completed": 0, "on_hold": 0, "employee_task_names": []
            }

    return {
        "total": frappe.db.count("Task Manager", filters),
        "open": frappe.db.count("Task Manager", {**filters, "status": "Open"}),
        "reopen": frappe.db.count("Task Manager", {**filters, "status": "Reopened"}),
        "in_progress": frappe.db.count("Task Manager", {**filters, "status": "In Progress"}),
        "completed": frappe.db.count("Task Manager", {**filters, "status": "Completed"}),
        "on_hold": frappe.db.count("Task Manager", {**filters, "status": "On Hold"}),
        "employee_task_names": tasks_list
    }


@frappe.whitelist()
def get_account_details(name):
    account = frappe.get_doc("Accounts", name).as_dict()
    if account.get("owner"):
        account["owner_name"] = frappe.db.get_value("User", account["owner"], "full_name") or account["owner"]
    return account



@frappe.whitelist()
def get_contacts_by_account(account_id, limit_start=0, limit_page_length=20):
    """
    Fetch contacts linked to a specific Account (company) via the Contact Company child table.
    This is a direct child-table lookup and works reliably regardless of filter handling.
    """
    import json
    limit_start = int(limit_start)
    limit_page_length = int(limit_page_length)

    # Find all Contact documents that have this account in their Contact Company child table
    linked_contacts = frappe.get_all(
        "Contact Company",
        filters={"company_name": account_id, "parenttype": "Contacts"},
        pluck="parent"
    )

    if not linked_contacts:
        return {"contacts": [], "total": 0}

    total = len(linked_contacts)
    paged_names = linked_contacts[limit_start: limit_start + limit_page_length]

    contacts = frappe.get_list(
        "Contacts",
        filters=[["Contacts", "name", "in", paged_names]],
        fields=[
            "name", "first_name", "email", "phone", "designation",
            "source_lead", "address", "notes", "country", "state",
            "city", "customer_type", "owner", "creation", "modified"
        ],
        order_by="creation desc"
    )

    # Enrich each contact with their company names
    contact_names = [c["name"] for c in contacts]
    child_entries = frappe.get_all(
        "Contact Company",
        filters={"parent": ["in", contact_names], "parenttype": "Contacts"},
        fields=["parent", "company_name"]
    )
    company_ids = list(set(e["company_name"] for e in child_entries if e.get("company_name")))
    account_map = {}
    if company_ids:
        accounts = frappe.get_all("Accounts", filters={"name": ["in", company_ids]}, fields=["name", "account_name"])
        account_map = {a["name"]: a["account_name"] for a in accounts}

    company_by_contact = {}
    for entry in child_entries:
        parent = entry["parent"]
        comp_id = entry["company_name"]
        comp_name = account_map.get(comp_id, comp_id)
        if comp_name:
            company_by_contact.setdefault(parent, []).append(comp_name)

    for c in contacts:
        c["company_names"] = company_by_contact.get(c["name"], [])
        c["company_name"] = ", ".join(c["company_names"])

    return {"contacts": contacts, "total": total}


def clean_contacts_or_filters(or_filters):
    cleaned_or_filters = []
    has_company_search = False
    company_search_val = ""
    for f in or_filters:
        if isinstance(f, list) and len(f) >= 3:
            # Field could be company_name
            fieldname = f[1] if len(f) > 3 or (len(f) == 3 and f[0] == "Contacts") else f[0]
            if fieldname == "company_name":
                has_company_search = True
                company_search_val = f[3] if len(f) > 3 else f[2]
            else:
                cleaned_or_filters.append(f)
        else:
            cleaned_or_filters.append(f)

    if has_company_search and company_search_val:
        matching_accounts = frappe.get_all("Accounts", filters={"account_name": ["like", company_search_val]}, pluck="name")
        if matching_accounts:
            matching_parents = frappe.get_all("Contact Company", filters={"company_name": ["in", matching_accounts]}, pluck="parent")
            if matching_parents:
                cleaned_or_filters.append(["Contacts", "name", "in", matching_parents])
            else:
                cleaned_or_filters.append(["Contacts", "name", "=", ""])
        else:
            cleaned_or_filters.append(["Contacts", "name", "=", ""])

    return cleaned_or_filters


@frappe.whitelist()
def get_contact_list(filters=None, or_filters=None, limit_start=0, limit_page_length=20, order_by="creation desc"):
    import json
    if isinstance(filters, str):
        filters = json.loads(filters)
    if isinstance(or_filters, str):
        or_filters = json.loads(or_filters)

    if or_filters:
        or_filters = clean_contacts_or_filters(or_filters)

    # Handle company_name filter in filters (AND filter) — it's a child table field,
    # so we resolve it manually and replace with name IN [...]
    if filters:
        new_filters = []
        for f in filters:
            if isinstance(f, list) and len(f) >= 3:
                fieldname = f[1] if len(f) == 4 else f[0]
                if fieldname == "company_name":
                    company_val = f[3] if len(f) == 4 else f[2]
                    # Lookup matching Account by name (exact)
                    matching_account = frappe.db.get_value("Accounts", {"account_name": company_val}, "name")
                    if matching_account:
                        matching_parents = frappe.get_all(
                            "Contact Company",
                            filters={"company_name": matching_account, "parenttype": "Contacts"},
                            pluck="parent"
                        )
                    else:
                        matching_parents = []
                    # Add as hard AND filter — contacts whose name is in matching_parents
                    new_filters.append(["Contacts", "name", "in", matching_parents if matching_parents else [""]])
                else:
                    new_filters.append(f)
            else:
                new_filters.append(f)
        filters = new_filters

    contacts = frappe.get_list(
        "Contacts",
        fields=[
            "name",
            "first_name",
            "email",
            "phone",
            "designation",
            "source_lead",
            "source_lead.lead_name",
            "address",
            "notes",
            "country",
            "state",
            "city",
            "customer_type",
            "owner",
            "creation",
            "modified"
        ],
        filters=filters,
        or_filters=or_filters,
        limit_start=limit_start,
        limit_page_length=limit_page_length,
        order_by=order_by
    )

    if not contacts:
        return []

    # Map lead_name to source_lead.lead_name if needed
    for c in contacts:
        if "lead_name" in c:
            c["source_lead.lead_name"] = c["lead_name"]

    # Fetch all child table entries for these contacts
    contact_names = [c["name"] for c in contacts]
    child_entries = frappe.get_all(
        "Contact Company",
        filters={"parent": ["in", contact_names], "parenttype": "Contacts"},
        fields=["parent", "company_name"]
    )

    # Fetch human-readable account names for these companies
    company_ids = list(set(e["company_name"] for e in child_entries if e.get("company_name")))
    account_map = {}
    if company_ids:
        accounts = frappe.get_all(
            "Accounts",
            filters={"name": ["in", company_ids]},
            fields=["name", "account_name"]
        )
        account_map = {a["name"]: a["account_name"] for a in accounts}

    # Group by parent
    company_by_contact = {}
    for entry in child_entries:
        parent = entry["parent"]
        comp_id = entry["company_name"]
        comp_name = account_map.get(comp_id, comp_id)
        if comp_name:
            if parent not in company_by_contact:
                company_by_contact[parent] = []
            company_by_contact[parent].append(comp_name)

    # Attach to contacts
    for c in contacts:
        c["company_names"] = company_by_contact.get(c["name"], [])
        c["company_name"] = ", ".join(c["company_names"])

    return contacts


@frappe.whitelist()
def get_contact_detail(name):
    doc = frappe.get_doc("Contacts", name)
    doc_dict = doc.as_dict()

    # Map company names to human-readable names
    company_ids = []
    if doc_dict.get("company_name"):
        company_ids = [row.get("company_name") for row in doc_dict["company_name"] if row.get("company_name")]
    
    account_map = {}
    if company_ids:
        accounts = frappe.get_all(
            "Accounts",
            filters={"name": ["in", company_ids]},
            fields=["name", "account_name"]
        )
        account_map = {a["name"]: a["account_name"] for a in accounts}

    # Format company_name for simple display
    company_names = [account_map.get(cid, cid) for cid in company_ids]
    doc_dict["company_name_display"] = ", ".join(company_names)
    doc_dict["company_name_list"] = company_names
    doc_dict["company_names"] = company_ids # keep raw IDs for form state

    return doc_dict


@frappe.whitelist()
def get_account_related_records(account_id, doctype, limit_start=0, limit_page_length=20):
    """
    Fetch related records for an Account.

    Relationships:
    - Deal        -> Directly linked using Deal.account
    - Invoice     -> Linked via Contacts (Invoice.client_name)
    - Estimation  -> Linked via Contacts (Estimation.client_name)
    - Purchase    -> Linked via Contacts (Purchase.client_name)
    """
    limit_start = int(limit_start)
    limit_page_length = int(limit_page_length)

    # Configuration for each supported doctype
    doctype_map = {
        "Invoice": {
            "fields": [
                "name", "ref_no", "client_name", "customer_name",
                "billing_name", "invoice_date", "grand_total",
                "received_amount", "balance_amount", "creation"
            ],
            "client_field": "client_name"
        },
        "Estimation": {
            "fields": [
                "name", "ref_no", "client_name", "customer_name",
                "estimate_date", "grand_total", "creation"
            ],
            "client_field": "client_name"
        },
        "Purchase": {
            "fields": [
                "name", "bill_no", "client_name", "bill_date",
                "grand_total", "paid_amount", "balance_amount", "creation"
            ],
            "client_field": "client_name"
        },
        "Deal": {
            "fields": [
                "name", "deal_title", "account", "contact",
                "stage", "value", "expected_close_date", "creation"
            ],
            # Deal is linked directly to Account
            "client_field": "account"
        }
    }

    if doctype not in doctype_map:
        frappe.throw(f"Unsupported doctype: {doctype}")

    config = doctype_map[doctype]

    # ---------------------------------------------------------
    # DEAL: Fetch directly by account (Contact is optional)
    # ---------------------------------------------------------
    if doctype == "Deal":
        all_records = frappe.get_list(
            "Deal",
            filters={"account": account_id},
            fields=config["fields"],
            order_by="creation desc",
            limit=None
        )

        total = len(all_records)
        paged = all_records[limit_start: limit_start + limit_page_length]

        return {
            "records": paged,
            "total": total
        }

    # ---------------------------------------------------------
    # OTHER DOCTYPES: Fetch via linked contacts
    # ---------------------------------------------------------
    linked_contact_ids = frappe.get_all(
        "Contact Company",
        filters={
            "company_name": account_id,
            "parenttype": "Contacts"
        },
        pluck="parent"
    )

    # If no contacts are linked, there can be no related
    # invoices/estimations/purchases.
    if not linked_contact_ids:
        return {
            "records": [],
            "total": 0
        }

    all_records = frappe.get_list(
        doctype,
        filters=[
            [doctype, config["client_field"], "in", linked_contact_ids]
        ],
        fields=config["fields"],
        order_by="creation desc",
        limit=None
    )

    total = len(all_records)
    paged = all_records[limit_start: limit_start + limit_page_length]

    return {
        "records": paged,
        "total": total
    }

@frappe.whitelist()
def get_followup_history(
    reference_type,
    reference_name
):
    """
    reference_type:
        Lead
        Contact
        Account

    reference_name:
        LEAD-0001
        CLT-0001
        CMP-0001
    """

    filters_calls = {}
    filters_meetings = {}

    reference_type = (
        reference_type or ""
    ).strip()

    reference_name = (
        reference_name or ""
    ).strip()

    if reference_type == "Lead":

        filters_calls["lead_name"] = (
            reference_name
        )

        filters_meetings["lead_name"] = (
            reference_name
        )

    elif reference_type == "Contact":

        filters_calls["contact_name"] = (
            reference_name
        )

        filters_meetings["contact_name"] = (
            reference_name
        )

    elif reference_type == "Account":

        filters_calls["account_name"] = (
            reference_name
        )

        filters_meetings["accounts_name"] = (
            reference_name
        )

    else:

        frappe.throw(
            "Invalid Reference Type"
        )

    # ----------------------------------------------------
    # CALLS
    # ----------------------------------------------------

    calls = frappe.get_all(
        "Calls",
        filters=filters_calls,
        fields=[
            "name",
            "title",
            "call_start_time",
            "call_end_time",
            "outgoing_call_status",
            "completed_call_status",
            "completed_call_notes",
            "owner_name",
            "creation"
        ]
    )

    # ----------------------------------------------------
    # MEETINGS
    # ----------------------------------------------------

    meetings = frappe.get_all(
        "Meeting",
        filters=filters_meetings,
        fields=[
            "name",
            "title",
            "from",
            "to",
            "outgoing_call_status",
            "completed_meet_status",
            "completed_meet_notes",
            "owner_name",
            "creation"
        ]
    )

    timeline = []

    # ----------------------------------------------------
    # CALL TIMELINE
    # ----------------------------------------------------

    for row in calls:

        timeline.append({

            "doctype":
                "Calls",

            "type":
                "Call",

            "name":
                row.name,

            "title":
                row.title,

            "status":
                row.outgoing_call_status,

            "result":
                row.completed_call_status,

            "notes":
                row.completed_call_notes,

            "owner":
                row.owner_name,

            "start_time":
                row.call_start_time,

            "end_time":
                row.call_end_time,

            "creation":
                row.creation

        })

    # ----------------------------------------------------
    # MEETING TIMELINE
    # ----------------------------------------------------

    for row in meetings:

        timeline.append({

            "doctype":
                "Meeting",

            "type":
                "Meeting",

            "name":
                row.name,

            "title":
                row.title,

            "status":
                row.outgoing_call_status,

            "result":
                row.completed_meet_status,

            "notes":
                row.completed_meet_notes,

            "owner":
                row.owner_name,

            "start_time":
                row.get("from"),

            "end_time":
                row.get("to"),

            "creation":
                row.creation

        })

    # ----------------------------------------------------
    # SORT TIMELINE
    # ----------------------------------------------------

    timeline.sort(
        key=lambda x:
            x.get("start_time")
            or x.get("creation"),
        reverse=True
    )

    return timeline

@frappe.whitelist()
def get_proposal_by_lead_id(lead_id):
    """
    Get all proposals linked to a Lead.

    Args:
        lead_id (str): Lead ID (e.g. LEAD-00342-2026)
    """

    if not lead_id:
        frappe.throw("Lead ID is required")

    proposals = frappe.get_all(
        "Proposal",
        filters={"lead": lead_id},
        fields=[
            "name",
            "proposal_title",
            "reference_no",
            "proposal_date",
            "valid_until",
            "status",
            "lead",
            "lead_name",
            "company_name",
            "creation",
            "modified"
        ],
        order_by="creation desc"
    )

    return proposals

# --------------------------------------------------------------------------
# LEAVE ALLOCATION
# --------------------------------------------------------------------------

def get_active_leave_types():
    return frappe.get_all(
        "Leave Type",
        filters={"status": "Active"},
        fields=[
            "name",
            "leave_type_name",
            "is_paid",
            "max_leaves",
            "carry_forward",
            "reset_frequency",
            "restrict_during_probation",
            "probation_period_months",
        ],
        order_by="creation asc",
    )

@frappe.whitelist()
def get_leave_allocation_preview(year: int, month: int):
    """
    Returns a preview of leave allocations for all active employees
    based on the Leave Type master.
    """

    year = int(year)
    month = int(month)

    month_start = get_first_day(datetime(year, month, 1))
    month_end = get_last_day(datetime(year, month, 1))

    prev_month_start = get_first_day(add_months(month_start, -1))
    prev_month_end = get_last_day(add_months(month_start, -1))

    freq_map = {
        "Every 3 months": 3,
        "Every 4 months": 4,
        "Every 6 months": 6,
        "Whole year": 12,
    }

    # Active Employees
    employees = frappe.get_all(
        "Employee",
        filters={"status": "Active"},
        fields=[
            "name",
            "employee_id",
            "employee_name",
            "date_of_joining",
            "skip_probation",
        ],
    )

    # Active Leave Types
    leave_types = get_active_leave_types()

    preview_data = []

    for emp in employees:

        allocations = []

        # ---------------------------------
        # Employee Probation Status (UI)
        # ---------------------------------
        in_probation = False

        if emp.date_of_joining and not emp.skip_probation:

            current_date = getdate(today())

            in_probation = any(
                add_months(
                    getdate(emp.date_of_joining),
                    lt.probation_period_months or 3
                ) >= current_date
                for lt in leave_types
                if lt.restrict_during_probation
            )

        for leave in leave_types:

            # ---------------------------------
            # Leave Type Probation Check
            # ---------------------------------
            leave_in_probation = False

            if (
                leave.restrict_during_probation
                and not emp.skip_probation
                and emp.date_of_joining
            ):
                probation_end = add_months(
                    getdate(emp.date_of_joining),
                    leave.probation_period_months or 3,
                )

                leave_in_probation = probation_end > month_end

            if leave_in_probation:
                continue

            # Already allocated?
            exists = frappe.db.exists(
                "Leave Allocation",
                {
                    "employee": emp.name,
                    "leave_type": leave.name,
                    "from_date": month_start,
                    "to_date": month_end,
                    "status": "Approved",
                },
            )

            carry_forward_balance = 0
            base_count = leave.max_leaves or 0

            # -----------------------------
            # Carry Forward Calculation
            # -----------------------------
            if leave.carry_forward:

                frequency = leave.reset_frequency or "Every 3 months"

                reset_interval = freq_map.get(
                    frequency,
                    3,
                )

                is_reset_month = (
                    (month - 1) % reset_interval
                ) == 0

                if not is_reset_month:

                    prev_alloc = frappe.get_value(
                        "Leave Allocation",
                        {
                            "employee": emp.name,
                            "leave_type": leave.name,
                            "from_date": prev_month_start,
                            "to_date": prev_month_end,
                            "status": "Approved",
                        },
                        [
                            "total_leaves_allocated",
                            "total_leaves_taken",
                        ],
                        as_dict=True,
                    )

                    if prev_alloc:

                        balance = (
                            flt(prev_alloc.total_leaves_allocated)
                            - flt(prev_alloc.total_leaves_taken)
                        )

                        if balance > 0:
                            carry_forward_balance = balance

            allocations.append({
                "leave_type": leave.name,
                "leave_type_name": leave.leave_type_name,
                "base_leaves": base_count,
                "carry_forward_balance": carry_forward_balance,
                "total_leaves": base_count + carry_forward_balance,
                "exists": bool(exists),
                "is_paid": leave.is_paid,
                "carry_forward": leave.carry_forward,
                "reset_frequency": leave.reset_frequency,
            })

        preview_data.append({
            "employee": emp.name,
            "employee_id": emp.employee_id,
            "employee_name": emp.employee_name,
            "date_of_joining": emp.date_of_joining,
            "in_probation": in_probation,
            "allocations": allocations,
        })

    return preview_data

@frappe.whitelist()
def auto_allocate_monthly_leaves(year: int, month: int):
    """
    Automatically allocate leaves for all active employees
    based on the Leave Type master.
    """

    year = int(year)
    month = int(month)

    try:
        month_start = get_first_day(datetime(year, month, 1))
        month_end = get_last_day(datetime(year, month, 1))

        prev_month_start = get_first_day(add_months(month_start, -1))
        prev_month_end = get_last_day(add_months(month_start, -1))

        # Active Employees
        employees = frappe.get_all(
            "Employee",
            filters={"status": "Active"},
            fields=[
                "name",
                "employee_id",
                "employee_name",
                "date_of_joining",
                "skip_probation",
            ],
        )

        # Active Leave Types
        leave_types = get_active_leave_types()

        freq_map = {
            "Every 3 months": 3,
            "Every 4 months": 4,
            "Every 6 months": 6,
            "Whole year": 12,
        }

        created_count = 0
        skipped_count = 0
        errors = []
        created_details = []

        for emp in employees:

            # ------------------------
            # Loop Through Leave Types
            # ------------------------
            for leave in leave_types:

                # Probation
                if (
                    leave.restrict_during_probation
                    and not emp.skip_probation
                    and emp.date_of_joining
                ):
                    probation_end = add_months(
                        getdate(emp.date_of_joining),
                        leave.probation_period_months or 3,
                    )

                    if probation_end > month_end:
                        continue

                leave_type = leave.leave_type_name
                base_leave_count = leave.max_leaves or 0

                try:

                    # Skip if already allocated
                    if frappe.db.exists(
                        "Leave Allocation",
                        {
                            "employee": emp.name,
                            "leave_type": leave.name,
                            "from_date": month_start,
                            "to_date": month_end,
                            "status": "Approved",
                        },
                    ):
                        skipped_count += 1
                        continue

                    # Previous Allocation
                    prev_alloc = frappe.get_value(
                        "Leave Allocation",
                        {
                            "employee": emp.name,
                            "leave_type": leave.name,
                            "from_date": prev_month_start,
                            "to_date": prev_month_end,
                            "status": "Approved",
                        },
                        [
                            "total_leaves_allocated",
                            "total_leaves_taken",
                        ],
                        as_dict=True,
                    )

                    carry_forward_balance = 0
                    leave_count = base_leave_count

                    # ------------------------
                    # Carry Forward Logic
                    # ------------------------
                    if leave.carry_forward:

                        frequency = (
                            leave.reset_frequency
                            or "Every 3 months"
                        )

                        reset_interval = freq_map.get(
                            frequency,
                            3,
                        )

                        is_reset_month = (
                            (month - 1) % reset_interval
                        ) == 0

                        if is_reset_month:
                            carry_forward_balance = 0

                        elif prev_alloc:

                            balance = (
                                flt(
                                    prev_alloc.total_leaves_allocated
                                )
                                - flt(
                                    prev_alloc.total_leaves_taken
                                )
                            )

                            if balance > 0:
                                carry_forward_balance = balance

                    total_allocation = (
                        leave_count + carry_forward_balance
                    )

                    # ------------------------
                    # Create Allocation
                    # ------------------------
                    allocation = frappe.get_doc(
                        {
                            "doctype": "Leave Allocation",
                            "employee": emp.name,
                            "leave_type": leave.name,
                            "from_date": month_start,
                            "to_date": month_end,
                            "total_leaves_allocated": total_allocation,
                            "total_leaves_taken": 0,
                            "status": "Approved",
                        }
                    )

                    allocation.insert(
                        ignore_permissions=True,
                        ignore_mandatory=True,
                    )

                    created_count += 1

                    created_details.append(
                        {
                            "employee_name": emp.employee_name,
                            "employee_id": emp.employee_id,
                            "leave_type": leave.name,
                            "allocated": total_allocation,
                            "carry_forward": carry_forward_balance,
                        }
                    )

                except Exception as e:
                    errors.append(
                        f"{emp.employee_id} - {leave_type} - {str(e)}"
                    )

        frappe.db.commit()

        return {
            "created_count": created_count,
            "skipped_count": skipped_count,
            "created_details": created_details,
            "errors": errors,
        }

    except Exception as e:
        frappe.throw(f"Error in auto leave allocation: {e}")


@frappe.whitelist()
def get_employee_probation_info(employee, date=None):
    """
    Returns probation status and restricted leave types for an employee.
    """

    if not employee:
        return {
            "is_probation": False,
            "restricted_types": [],
            "probation_end_date": None,
        }

    emp = frappe.db.get_value(
        "Employee",
        employee,
        ["date_of_joining", "skip_probation"],
        as_dict=True,
    )

    if (
        not emp
        or not emp.date_of_joining
        or emp.skip_probation
    ):
        return {
            "is_probation": False,
            "restricted_types": [],
            "probation_end_date": None,
        }

    check_date = getdate(date) if date else getdate(today())

    leave_types = frappe.get_all(
        "Leave Type",
        filters={
            "status": "Active",
            "restrict_during_probation": 1,
        },
        fields=[
            "name",
            "probation_period_months",
        ],
    )

    restricted_types = []
    probation_end_date = None

    for leave in leave_types:

        months = leave.probation_period_months or 3

        end_date = add_months(
            getdate(emp.date_of_joining),
            months,
        )

        if end_date >= check_date:
            restricted_types.append(leave.name)

            if (
                probation_end_date is None
                or end_date > probation_end_date
            ):
                probation_end_date = end_date

    return {
        "is_probation": bool(restricted_types),
        "restricted_types": restricted_types,
        "probation_end_date": probation_end_date,
    }

@frappe.whitelist()
def get_automation_options():
    """
    Get necessary options for the CRM WhatsApp Automation frontend.
    Returns Lead workflow states, active workflow name, Deal stages, and Lead fields.
    """
    options = {
        "lead_workflow_states": [],
        "deal_stages": [],
        "lead_fields": [],
        "active_lead_workflow": None
    }
    
    # 1. Active Lead Workflow and States
    workflow = frappe.db.get_value("Workflow", {"document_type": "Lead", "is_active": 1}, "name")
    if workflow:
        options["active_lead_workflow"] = workflow
        states = frappe.get_all(
            "Workflow Document State",
            filters={"parent": workflow},
            fields=["state"],
            order_by="idx"
        )
        options["lead_workflow_states"] = [s.state for s in states]
        
    # 2. Deal Stages
    try:
        deal_meta = frappe.get_meta("Deal")
        stage_field = deal_meta.get_field("stage")
        if stage_field and stage_field.options:
            options["deal_stages"] = [s for s in stage_field.options.split("\n") if s.strip()]
    except Exception:
        pass
        
    # 3. Lead Fields
    try:
        lead_meta = frappe.get_meta("Lead")
        for df in lead_meta.fields:
            if df.fieldtype not in ("Section Break", "Column Break", "Tab Break", "HTML", "Button"):
                options["lead_fields"].append({
                    "fieldname": df.fieldname,
                    "label": df.label,
                    "fieldtype": df.fieldtype
                })
    except Exception:
        pass
        
    return options
